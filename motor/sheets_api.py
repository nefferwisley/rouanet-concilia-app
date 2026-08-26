"""Portas de planilha e adaptadores em lote para o espelho financeiro."""

from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Protocol, Sequence

import openpyxl

from motor.conflito import VersãoRegistro


class SheetsPort(Protocol):
    """Contrato mínimo; a implementação OAuth real não pertence ao domínio."""

    def ler_lote(self, projeto_id: str) -> list[VersãoRegistro]: ...

    def escrever_lote(
        self,
        projeto_id: str,
        registros: Sequence[VersãoRegistro],
    ) -> None: ...


class SheetsMemória:
    """Adaptador determinístico para testes e desenvolvimento sem egresso."""

    def __init__(self) -> None:
        self._projetos: dict[str, dict[str, VersãoRegistro]] = {}
        self.escritas_em_lote = 0

    def ler_lote(self, projeto_id: str) -> list[VersãoRegistro]:
        return list(self._projetos.get(projeto_id, {}).values())

    def escrever_lote(
        self,
        projeto_id: str,
        registros: Sequence[VersãoRegistro],
    ) -> None:
        if not registros:
            return
        destino = self._projetos.setdefault(projeto_id, {})
        for registro in registros:
            destino[registro.registro_id] = registro
        self.escritas_em_lote += 1

    def simular_edição(self, projeto_id: str, registro: VersãoRegistro) -> None:
        """Representa uma edição manual recebida da planilha."""
        self._projetos.setdefault(projeto_id, {})[registro.registro_id] = registro


ABA_REVISADA = "CONCILIAÇÃO REVISADA"

# Contrato explícito entre os conceitos canônicos e os cabeçalhos aceitos.
# A posição não faz parte do contrato: usuários podem inserir/reordenar colunas.
COLUNAS_ESPELHO: dict[str, tuple[str, ...]] = {
    "controle": ("CONTROLE",),
    "entrada": ("ENTRADA",),
    "valor_entrada": ("VALOR ENTRADA",),
    "prestador": ("PRESTADOR DE SERVIÇO", "PRESTADOR DE SERVICO", "PRESTADOR"),
    "razao_social": ("RAZÃO SOCIAL", "RAZAO SOCIAL"),
    "data": ("DATA", "DATA DE PAGAMENTO"),
    "valor": ("VALOR", "VALOR PAGO"),
    "saldo": ("SALDO",),
    "item": ("ITEM",),
    "rubrica": ("RUBRICA", "RUBRICA SALIC"),
    "status_revisao": ("STATUS DA REVISÃO", "STATUS DA REVISAO"),
    "documento_fiscal": ("DOCUMENTO FISCAL", "DOC FISCAL"),
    "evidencia": ("PRINT (EVIDÊNCIA)", "PRINT (EVIDENCIA)"),
}

METADADOS_SYNC = {
    "_sync_id": "_SYNC_ID",
    "_sync_version": "_SYNC_VERSION",
    "_sync_updated_by": "_SYNC_UPDATED_BY",
    "_sync_updated_at": "_SYNC_UPDATED_AT",
}


def _normalizar_cabeçalho(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().upper()


def _mapear_colunas(ws, limite: int = 20) -> tuple[int, dict[str, int]]:
    """Localiza o cabeçalho e devolve índices 1-based por conceito."""
    aliases = {
        conceito: {_normalizar_cabeçalho(nome) for nome in nomes}
        for conceito, nomes in COLUNAS_ESPELHO.items()
    }
    aliases.update(
        {conceito: {_normalizar_cabeçalho(nome)} for conceito, nome in METADADOS_SYNC.items()}
    )
    for linha in range(1, min(ws.max_row or 1, limite) + 1):
        existentes = {
            _normalizar_cabeçalho(ws.cell(linha, coluna).value): coluna
            for coluna in range(1, (ws.max_column or 1) + 1)
            if ws.cell(linha, coluna).value not in (None, "")
        }
        mapeadas = {
            conceito: existentes[nome]
            for conceito, nomes in aliases.items()
            for nome in nomes
            if nome in existentes
        }
        if {"prestador", "data", "valor"} <= mapeadas.keys():
            return linha, mapeadas
    raise ValueError("Cabeçalho PRESTADOR/DATA/VALOR não encontrado na aba revisada.")


def _decimal(valor: Any) -> Decimal | None:
    if valor in (None, "") or (isinstance(valor, str) and valor.startswith("=")):
        return None
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _texto(valor: Any) -> str | None:
    if valor in (None, "") or (isinstance(valor, str) and valor.startswith("=")):
        return None
    if isinstance(valor, (int, float, Decimal)) and not isinstance(valor, bool):
        número = Decimal(str(valor))
        return format(número.normalize(), "f")
    return str(valor).strip() or None


def _data(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto[:10], formato).date()
        except ValueError:
            pass
    return None


class SheetsXlsx:
    """Espelho local seguro: lê o modelo e grava sempre em uma cópia separada."""

    def __init__(
        self,
        modelo: Path,
        saída: Path,
        aba: str = ABA_REVISADA,
    ) -> None:
        self.modelo = Path(modelo)
        self.saída = Path(saída)
        self.aba = aba
        if self.modelo.resolve() == self.saída.resolve():
            raise ValueError("A saída do espelho não pode sobrescrever o arquivo-modelo.")

    @property
    def caminho_leitura(self) -> Path:
        return self.saída if self.saída.exists() else self.modelo

    def _abrir(self, *, data_only: bool = False):
        if not self.caminho_leitura.exists():
            raise FileNotFoundError(self.caminho_leitura)
        return openpyxl.load_workbook(self.caminho_leitura, data_only=data_only)

    def ler_lote(self, projeto_id: str) -> list[VersãoRegistro]:
        del projeto_id  # Um arquivo representa a projeção de um único projeto.
        wb = self._abrir(data_only=True)
        if self.aba not in wb.sheetnames:
            raise ValueError(f"Aba obrigatória ausente: {self.aba}")
        ws = wb[self.aba]
        cabeçalho, colunas = _mapear_colunas(ws)
        registros: list[VersãoRegistro] = []

        for linha in range(cabeçalho + 1, (ws.max_row or cabeçalho) + 1):
            data = _data(ws.cell(linha, colunas["data"]).value)
            valor = _decimal(ws.cell(linha, colunas["valor"]).value)
            if data is None or valor is None:
                continue

            def célula(conceito: str) -> Any:
                coluna = colunas.get(conceito)
                return ws.cell(linha, coluna).value if coluna else None

            controle = _texto(célula("controle"))
            registro_id = str(célula("_sync_id") or "").strip()
            if not registro_id:
                controle_id = str(controle or "").strip()
                registro_id = f"controle:{controle_id}" if controle_id else f"linha:{linha}"

            versão_raw = célula("_sync_version")
            try:
                versão = int(versão_raw or 0)
            except (TypeError, ValueError):
                versão = 0
            atualizado_em_raw = célula("_sync_updated_at")
            atualizado_em = (
                atualizado_em_raw
                if isinstance(atualizado_em_raw, datetime)
                else datetime.now(timezone.utc)
            )
            campos = {
                "controle": controle,
                "entrada": _texto(célula("entrada")),
                "valor_entrada": _decimal(célula("valor_entrada")),
                "prestador": _texto(célula("prestador")),
                "razao_social": _texto(célula("razao_social")),
                "data": data,
                "valor": valor,
                # SALDO é projeção calculada por fórmula no XLSX. Ele não entra
                # no estado canônico nem no hash de conflito, pois o openpyxl
                # preserva a fórmula mas não recalcula seu valor em cache.
                "saldo": None,
                "item": _texto(célula("item")),
                "rubrica": _texto(célula("rubrica")),
                "status_revisao": _texto(célula("status_revisao")),
                "documento_fiscal": _texto(célula("documento_fiscal")),
                "evidencia": _texto(célula("evidencia")),
            }
            registros.append(
                VersãoRegistro(
                    registro_id=registro_id,
                    versão=versão,
                    campos=campos,
                    atualizado_por=str(célula("_sync_updated_by") or "planilha"),
                    atualizado_em=atualizado_em,
                )
            )
        wb.close()
        return registros

    def escrever_lote(
        self,
        projeto_id: str,
        registros: Sequence[VersãoRegistro],
    ) -> None:
        del projeto_id
        if not registros:
            return
        if not self.saída.exists():
            self.saída.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(self.modelo, self.saída)

        wb = openpyxl.load_workbook(self.saída, data_only=False)
        if self.aba not in wb.sheetnames:
            raise ValueError(f"Aba obrigatória ausente: {self.aba}")
        ws = wb[self.aba]
        cabeçalho, colunas = _mapear_colunas(ws)

        for conceito, nome in METADADOS_SYNC.items():
            if conceito not in colunas:
                coluna = (ws.max_column or 0) + 1
                ws.cell(cabeçalho, coluna, nome)
                ws.column_dimensions[ws.cell(cabeçalho, coluna).column_letter].hidden = True
                colunas[conceito] = coluna

        linhas_por_id: dict[str, int] = {}
        for linha in range(cabeçalho + 1, (ws.max_row or cabeçalho) + 1):
            sync_id = ws.cell(linha, colunas["_sync_id"]).value
            if sync_id:
                linhas_por_id[str(sync_id)] = linha
            else:
                controle = ws.cell(linha, colunas.get("controle", 1)).value
                if controle not in (None, "") and not (
                    isinstance(controle, str) and controle.startswith("=")
                ):
                    controle_id = _texto(controle)
                    if controle_id:
                        linhas_por_id[f"controle:{controle_id}"] = linha
                elif (
                    ws.cell(linha, colunas["data"]).value not in (None, "")
                    and ws.cell(linha, colunas["valor"]).value not in (None, "")
                ):
                    linhas_por_id[f"linha:{linha}"] = linha

        # CONTROLE/ENTRADA/VALOR ENTRADA/SALDO podem ser fórmulas estruturais
        # do modelo e nunca são sobrescritos pelo espelho.
        graváveis = {
            "prestador",
            "razao_social",
            "data",
            "valor",
            "item",
            "rubrica",
            "status_revisao",
            "documento_fiscal",
            "evidencia",
        }
        for registro in registros:
            linha = linhas_por_id.get(registro.registro_id)
            if linha is None:
                linha = (ws.max_row or cabeçalho) + 1
            for conceito in graváveis:
                coluna = colunas.get(conceito)
                if coluna and conceito in registro.campos:
                    ws.cell(linha, coluna, registro.campos[conceito])
            ws.cell(linha, colunas["_sync_id"], registro.registro_id)
            ws.cell(linha, colunas["_sync_version"], registro.versão)
            ws.cell(linha, colunas["_sync_updated_by"], registro.atualizado_por)
            ws.cell(linha, colunas["_sync_updated_at"], registro.atualizado_em.replace(tzinfo=None))

        wb.save(self.saída)
        wb.close()

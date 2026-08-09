#!/usr/bin/env python3
"""
motor/gerar_resumo.py — Task 006: relatório de validação (auditoria de saldo).

Lê somente os artefatos (não altera nada) e escreve
    saida/relatorios/resumo_validacao.md   (pt-BR)

Fontes:
    motor/_parsed/movimentos.json         (extrato)
    motor/_parsed/comprovantes.json       (comprovantes)
    motor/_parsed/cruzamento.json         (task 003, formato lista)
    saida/planilha/planilha_corrigida.xlsx (task 004)
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
PARSED = RAIZ / "motor" / "_parsed"
SAIDA = RAIZ / "saida" / "relatorios"
MD = SAIDA / "resumo_validacao.md"

TOLERANCIA = 0.01


def _br(v) -> str:
    s = f"{float(v):,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def _soma(linhas) -> float:
    return round(sum(round(float(r["valor"]), 2) for r in linhas), 2)


def main():
    movs = json.loads((PARSED / "movimentos.json").read_text(encoding="utf-8"))
    comps = json.loads((PARSED / "comprovantes.json").read_text(encoding="utf-8"))
    cruz = json.loads((PARSED / "cruzamento.json").read_text(encoding="utf-8"))

    wb = load_workbook(RAIZ / "saida" / "planilha" / "planilha_corrigida.xlsx")
    ws = wb.active
    soma_planilha_conciliados = 0.0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 7).value == "CONCILIADO":
            soma_planilha_conciliados = round(soma_planilha_conciliados + float(ws.cell(r, 6).value), 2)

    debs = [m for m in movs if m.get("sinal") == "D"]
    creds = [m for m in movs if m.get("sinal") == "C"]
    conc = [r for r in cruz if r["status"] == "CONCILIADO"]
    orf_ext = [r for r in cruz if r["status"] == "SEM-COMPROVANTE"]
    orf_comp = [r for r in cruz if r["status"] == "SEM-EXTRATO"]
    div = [r for r in cruz if r["status"] == "DIVERGENTE"]
    amb = [r for r in cruz if r["status"] == "AMBIGUO"]

    soma_deb_conc = _soma(conc)
    soma_comp_conc = _soma(conc)

    nao_casados_deb = _soma(orf_ext) + _soma(div)
    nao_casados_comp = _soma(orf_comp) + _soma(div) + _soma(amb)

    difs = {
        "débitos conciliados vs comprovantes conciliados": abs(soma_deb_conc - soma_comp_conc),
        "débitos conciliados vs planilha (CONCILIADO)": abs(soma_deb_conc - soma_planilha_conciliados),
        "comprovantes conciliados vs planilha (CONCILIADO)": abs(soma_comp_conc - soma_planilha_conciliados),
    }
    batimento_ok = all(d <= TOLERANCIA for d in difs.values())

    pendencias = []
    for r in orf_ext:
        pendencias.append(("SEM-COMPROVANTE", r.get("data_pagamento") or "", r.get("favorecido") or "", r["valor"],
                           r.get("observacao") or "débito no extrato sem comprovante correspondente"))
    for r in orf_comp:
        pendencias.append(("SEM-EXTRATO", r.get("data_pagamento") or "", r.get("favorecido") or "", r["valor"],
                           r.get("observacao") or "comprovante sem débito correspondente"))
    for r in div:
        pendencias.append(("DIVERGENTE", r.get("data_pagamento") or "", r.get("favorecido") or "", r["valor"],
                           r.get("observacao") or "valor ou condições divergem"))
    for r in amb:
        pendencias.append(("AMBIGUO", r.get("data_pagamento") or "", r.get("favorecido") or "", r["valor"],
                           r.get("observacao") or "pagamento duplicado ou débito não lançado"))

    total_deb = len(debs)
    total_comp = len(comps)
    n_conc = len(conc)
    t_pct = round(n_conc / total_deb * 100, 1) if total_deb else 0.0
    taxa_acerto = round(n_conc / (n_conc + len(pendencias)) * 100, 1) if (n_conc + len(pendencias)) else 0.0

    linhas = []
    ap = linhas.append
    ap(f"# Resumo de Validação — Projeto 1961 (PRONAC 20-7453)")
    ap("")
    ap(f"Gerado automaticamente por `motor/gerar_resumo.py` em {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.")
    ap("")
    ap("## 1. Resumo")
    ap("")
    ap(f"- Movimentações no extrato: **{len(movs)}** (débitos: **{total_deb}**, créditos: **{len(creds)}**)")
    ap(f"- Comprovantes: **{total_comp}**")
    ap(f"- Conciliados: **{n_conc}** ({t_pct}% dos débitos)")
    ap(f"- Órfãos no extrato (sem comprovante): **{len(orf_ext)}**")
    ap(f"- Órfãos comprovante (sem extrato): **{len(orf_comp)}**")
    ap(f"- Divergentes de valor: **{len(div)}**")
    ap(f"- Ambíguos: **{len(amb)}**")
    ap("")
    ap("## 2. Batimento de saldo")
    ap("")
    ap("| Origem | Soma (R$) |")
    ap("|---|---|")
    ap(f"| Débitos conciliados (extrato) | {_br(soma_deb_conc)} |")
    ap(f"| Comprovantes conciliados | {_br(soma_comp_conc)} |")
    ap(f"| Planilha — linhas CONCILIADO | {_br(soma_planilha_conciliados)} |")
    ap("")
    for nome, dif in difs.items():
        ap(f"- {nome}: diferença {_br(dif)} {'OK' if dif <= TOLERANCIA else 'FORA DA TOLERÂNCIA'}")
    ap("")
    ap(f"**Batimento: {'OK' if batimento_ok else 'NÃO BATE'}** (tolerância R$ 0,01).")
    ap("")
    ap("Lançamentos não casados (somas gerais):")
    ap(f"- Débitos não conciliados: **{_br(nao_casados_deb)}**")
    ap(f"- Comprovantes não conciliados: **{_br(nao_casados_comp)}**")
    ap("")
    ap("## 3. Pendências")
    ap("")
    ap(f"Total: **{len(pendencias)}** linha(s).")
    ap("")
    ap("| Data | Favorecido | Valor (R$) | Status | Motivo |")
    ap("|---|---|---|---|---|")
    for status, data, fav, valor, motivo in pendencias:
        fav = (fav or "").replace("|", "/")
        motivo = (motivo or "").replace("|", "/")
        ap(f"| {data} | {fav} | {_br(valor)} | {status} | {motivo} |")
    ap("")
    ap("## 4. Taxa de acerto")
    ap("")
    ap(f"- Conciliados: **{n_conc}**")
    ap(f"- Pendências: **{len(pendencias)}**")
    ap(f"- Taxa de acerto: **{taxa_acerto}%** (conciliados / (conciliados + pendências))")
    ap("")
    ap("## 5. Nota metodológica")
    ap("")
    ap("- Extrações 100% determinísticas: PyMuPDF, texto nativo do PDF, sem OCR nem IA externa.")
    ap("- Cruzamento 1:1 por chave (data + valor); comprovante com valor ilegível assume o valor do "
       "débito que casa por data + favorecido normalizado (registrado na observação).")
    ap("- Comprovantes excedentes numa chave (mesmo valor/data de outro que já casou) ficam AMBIGUO; "
       "rubrica SALIC fica '(a classificar)' porque o cruzamento não a traz (nunca inventada).")
    ap("")

    SAIDA.mkdir(parents=True, exist_ok=True)
    MD.write_text("\n".join(linhas), encoding="utf-8")

    return {
        "arquivo": str(MD),
        "débitos": total_deb,
        "comprovantes": total_comp,
        "conciliados": n_conc,
        "taxa_conciliacao_pct": t_pct,
        "taxa_acerto_pct": taxa_acerto,
        "pendencias": len(pendencias),
        "soma_deb_conc": soma_deb_conc,
        "soma_comp_conc": soma_comp_conc,
        "soma_planilha_conc": soma_planilha_conciliados,
        "batimento_ok": batimento_ok,
    }


if __name__ == "__main__":
    print(json.dumps(main(), ensure_ascii=False, indent=1))

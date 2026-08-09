#!/usr/bin/env python3
"""
motor/gerar_planilha.py — Task 004: planilha corrigida (xlsx) a partir do cruzamento.

Entrada:
    motor/_parsed/cruzamento.json   (task 003)

Saídas:
    saida/planilha/planilha_corrigida.xlsx
    motor/_parsed/planilha_linhas.json   (referência p/ a task 005)

Uma linha = um pagamento (conciliados + órfãos + divergentes + ambíguos).
O cruzamento não traz rubrica SALIC -> coluna fica "(a classificar)" (nunca
inventar; contagem reportada ao board).

planilha_linhas.json só lista as linhas QUE TÊM comprovante (arquivo final),
na mesma ordem das linhas da planilha — é o conjunto que a task 005 copia.
Linhas SEM-COMPROVANTE ficam com a coluna "Arquivo Final" vazia.
Comprovante com valor ilegível (<= 0) -> arquivo final = nome original em
"saida/arquivos_finais/PENDENTES/" (regra da task 005).

Colunas (ordem fixa): Nº, Data pagamento, Favorecido, CNPJ/CPF, Rubrica SALIC,
Valor, Status, Arquivo Final, Observação.
"""

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
PARSED = RAIZ / "motor" / "_parsed"
PASTA_PLANILHA = RAIZ / "saida" / "planilha"
XLSX = PASTA_PLANILHA / "planilha_corrigida.xlsx"
LINHAS_JSON = PARSED / "planilha_linhas.json"

RUBRICA = "(a classificar)"
PENDENTES = "PENDENTES"

# classe (task 003) -> status (vocabulário da planilha)
CLASSE_PARA_STATUS = {
    "conciliados": "CONCILIADO",
    "orfaos_extrato": "SEM-COMPROVANTE",
    "orfaos_comprovante": "SEM-EXTRATO",
    "divergentes_valor": "DIVERGENTE",
    "ambiguos_extrato": "AMBIGUO",
    "ambiguos_comprovante": "AMBIGUO",
}

ORDEM = [
    "conciliados",
    "ambiguos_extrato",
    "ambiguos_comprovante",
    "divergentes_valor",
    "orfaos_extrato",
    "orfaos_comprovante",
]

CABECALHO = [
    "Nº", "Data pagamento", "Favorecido", "CNPJ/CPF", "Rubrica SALIC",
    "Valor", "Status", "Arquivo Final", "Observação",
]


# ---------------------------------------------------------------- helpers
def _normalizar(s) -> str:
    t = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in t if not unicodedata.combining(c))


def _slug(s) -> str:
    t = _normalizar(s).lower()
    t = re.sub(r"[^a-z0-9]+", "_", t)
    return t.strip("_")


def _data_ddmm(iso) -> str:
    a, m, d = str(iso).split("-")
    return f"{d}-{m}-{a}"


def _valor_br(v) -> str:
    s = f"{float(v):,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R${s}"


def _sem_sufixo_truncado(favorecido) -> str:
    return re.sub(r"\(\s*truncado\s*\)\s*$", "", favorecido).strip()


def _resumo_unicos(nomes) -> str:
    vistos = []
    for n in nomes:
        if n and n not in vistos:
            vistos.append(n)
    return ", ".join(vistos)


def _fav_debito(m) -> str:
    return (m.get("favorecido") or "").strip()


def _fav_comprovante(c) -> str:
    return (c.get("favorecido") or c.get("descricao") or "").strip()


def _valor_debito(m):
    return round(float(m["valor"]), 2)


def _valor_comprovante(c):
    v = c.get("valor")
    return round(float(v), 2) if v is not None else None


def _linhas_brutas(cruz) -> list[dict]:
    """Constrói uma linha (sem Nº/Arquivo Final) por item do cruzamento."""
    linhas = []
    for classe in ORDEM:
        status = CLASSE_PARA_STATUS[classe]
        for item in cruz.get(classe, []):
            if classe == "conciliados":
                d, c = item["debito"], item["comprovante"]
                linhas.append({
                    "data": d["data"],
                    "favorecido": _fav_comprovante(c) or _fav_debito(d),
                    "cnpj": c.get("cnpj"),
                    "valor": _valor_comprovante(c) if _valor_comprovante(c) is not None else _valor_debito(d),
                    "status": status,
                    "obs": "",
                    "numero_arquivo": c.get("numero_arquivo"),
                    "fonte_comprovante": c.get("fonte"),
                    "valor_comprovante": _valor_comprovante(c),
                })
            elif classe == "ambiguos_extrato":
                d = item["debito"]
                linhas.append({
                    "data": d["data"],
                    "favorecido": _fav_debito(d) + " (truncado)",
                    "cnpj": None,
                    "valor": _valor_debito(d),
                    "status": status,
                    "obs": "débito disputado por comprovantes: " + _resumo_unicos(item.get("candidatos_comprovantes", [])),
                    "numero_arquivo": None,
                    "fonte_comprovante": None,
                    "valor_comprovante": None,
                })
            elif classe == "ambiguos_comprovante":
                c = item["comprovante"]
                linhas.append({
                    "data": c["data"],
                    "favorecido": _fav_comprovante(c),
                    "cnpj": c.get("cnpj"),
                    "valor": _valor_comprovante(c) if _valor_comprovante(c) is not None else 0.0,
                    "status": status,
                    "obs": "comprovante disputado por débitos: " + _resumo_unicos(item.get("candidatos_extrato", [])),
                    "numero_arquivo": c.get("numero_arquivo"),
                    "fonte_comprovante": c.get("fonte"),
                    "valor_comprovante": _valor_comprovante(c),
                })
            elif classe == "divergentes_valor":
                d, c = item["debito"], item["comprovante"]
                linhas.append({
                    "data": d["data"],
                    "favorecido": _fav_comprovante(c) or _fav_debito(d),
                    "cnpj": c.get("cnpj"),
                    "valor": _valor_debito(d),
                    "status": status,
                    "obs": item.get("motivo", ""),
                    "numero_arquivo": c.get("numero_arquivo"),
                    "fonte_comprovante": c.get("fonte"),
                    "valor_comprovante": _valor_comprovante(c),
                })
            elif classe == "orfaos_extrato":
                d = item["debito"]
                linhas.append({
                    "data": d["data"],
                    "favorecido": _fav_debito(d) + " (truncado)",
                    "cnpj": None,
                    "valor": _valor_debito(d),
                    "status": status,
                    "obs": item.get("observacao", ""),
                    "numero_arquivo": None,
                    "fonte_comprovante": None,
                    "valor_comprovante": None,
                })
            else:  # orfaos_comprovante
                c = item["comprovante"]
                linhas.append({
                    "data": c["data"],
                    "favorecido": _fav_comprovante(c),
                    "cnpj": c.get("cnpj"),
                    "valor": _valor_comprovante(c) if _valor_comprovante(c) is not None else 0.0,
                    "status": status,
                    "obs": item.get("observacao", ""),
                    "numero_arquivo": c.get("numero_arquivo"),
                    "fonte_comprovante": c.get("fonte"),
                    "valor_comprovante": _valor_comprovante(c),
                })
    return linhas


def _arquivo_final(num, data, valor, favorecido):
    slug_fav = _slug(_sem_sufixo_truncado(favorecido)) or "sem_favorecido"
    return f"{num:04d}_{RUBRICA}_{_data_ddmm(data)}_{_valor_br(valor)}_{slug_fav}.pdf"


def _destino_linha(ln):
    """Subpasta + arquivo final de cada linha (None se a linha não tem comprovante)."""
    if ln["numero_arquivo"] is None:
        return None, None
    if ln["valor_comprovante"] is None or ln["valor_comprovante"] <= 0:
        return PENDENTES, ln["fonte_comprovante"]
    return RUBRICA, _arquivo_final(ln["num"], ln["data"], ln["valor"], ln["favorecido"])


def _gerar_xlsx(linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    ws.append(CABECALHO)
    for col, _ in enumerate(CABECALHO, start=1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="DDEBF7")
        ws.cell(row=1, column=col).alignment = Alignment(vertical="center")

    for ln in linhas:
        arq = ln["arquivo_final"]
        ws.append([
            ln["num"],
            ln["data"],
            ln["favorecido"],
            ln["cnpj"] or "",
            RUBRICA,
            ln["valor"],
            ln["status"],
            (f"{ln['subpasta']}\\{arq}" if arq else ""),
            ln["obs"],
        ])
        ws.cell(row=ws.max_row, column=6).number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    larguras = [6, 13, 42, 20, 17, 13, 16, 62, 60]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    PASTA_PLANILHA.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)


def main():
    cruz = json.loads((PARSED / "cruzamento.json").read_text(encoding="utf-8"))
    brutas = _linhas_brutas(cruz)

    usados = Counter()
    linhas = []
    for n, ln in enumerate(brutas, start=1):
        ln = {**ln, "num": n}
        subpasta, arq = _destino_linha(ln)
        if arq:
            usados[arq] += 1
            if usados[arq] > 1:
                arq = f"{arq[:-4]}_{usados[arq]}.pdf"
        linhas.append({**ln, "subpasta": subpasta, "arquivo_final": arq})

    _gerar_xlsx(linhas)

    ref = [
        {
            "arquivo_final": ln["arquivo_final"],
            "subpasta": ln["subpasta"],
            "data": ln["data"],
            "valor": ln["valor"],
            "numero_arquivo": ln["numero_arquivo"],
        }
        for ln in linhas
        if ln["arquivo_final"]
    ]
    LINHAS_JSON.write_text(json.dumps(ref, ensure_ascii=False, indent=1), encoding="utf-8")

    return {
        "arquivo": str(XLSX),
        "linhas": len(linhas),
        "com_arquivo_final": len(ref),
        "por_status": dict(Counter(ln["status"] for ln in linhas)),
        "rubrica": RUBRICA,
        "pendentes": sum(1 for ln in linhas if ln["subpasta"] == PENDENTES),
    }


if __name__ == "__main__":
    print(json.dumps(main(), ensure_ascii=False, indent=1))

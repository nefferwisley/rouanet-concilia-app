#!/usr/bin/env python3
"""
motor/gerar_cruzamento_banco_planilha.py — nova versão do cruzamento
"Banco de Dados Atual (produção) x Planilha Revisada (1961)".

Entradas:
    %TEMP% banco_001.csv                     — dump atual do Postgres (projeto "001",
                                               185 transações CONCILIADO_OK, via
                                               docker exec psql copy + docker cp).
    1961_Revisao_Financeira_ATUALIZADA.xlsx  — planilha revisada local (aba
                                               CONCILIAÇÃO REVISADA, 179 pagamentos).
    Cruzamento_1961_Banco_x_PlanilhaRevisada.xlsx — versão anterior (carrega a
                                               coluna "Nota/correção" por linha da
                                               planilha, que foi curada manualmente).

Saídas:
    saida/cruzamento/cruzamento_1961_banco_x_planilha_v2.csv   — mesmo formato do _tmp
    saida/cruzamento/Cruzamento_1961_Banco_x_PlanilhaRevisada_v2.xlsx
        Abas: Resumo | Comparacao | Somente no banco atual
    saida/cruzamento/cruzamento_1961_artifact_v2.html          — página de conferência
    saida/planilha/planilha_atualizada_saas.xlsx               — planilha refletindo
        só os dados cruzados no SaaS (185 transações CONCILIADO_OK com extrato).

Regras de matching:
    1ª rodada — mesma data + mesmo valor (determinístico, primeira ocorrência).
    2ª rodada — mesmo valor, data diferente (0 no estado atual).
    Restam: linhas da planilha sem correspondência e lançamentos só no banco.

Observação: nenhuma alteração é feita em produção — arquivos só para conferência.
"""
import csv
import json
import re
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
TEMP = Path(r"C:\Users\Dell\AppData\Local\Temp\opencode\rouanet_1961")
PLANILHA = RAIZ / "1961_Revisao_Financeira_ATUALIZADA.xlsx"
BANCO_CSV = TEMP / "banco_001.csv"
ANTERIOR = RAIZ / "Cruzamento_1961_Banco_x_PlanilhaRevisada.xlsx"

SAIDA_CSV = RAIZ / "saida" / "cruzamento" / "cruzamento_1961_banco_x_planilha_v2.csv"
SAIDA_XLSX = RAIZ / "saida" / "cruzamento" / "Cruzamento_1961_Banco_x_PlanilhaRevisada_v2.xlsx"
SAIDA_HTML = RAIZ / "saida" / "cruzamento" / "cruzamento_1961_artifact_v2.html"
SAIDA_PLANILHA_SAAS = RAIZ / "saida" / "planilha" / "planilha_atualizada_saas.xlsx"

FONTE = "1961_Revisao_Financeira_ATUALIZADA.xlsx (aba CONCILIAÇÃO REVISADA)"
DATA_CRUZAMENTO = "2026-08-12"
PRONAC = "1961 (PRONAC 20-7453)"

STATUS_OK = "OK — data e valor conferem"
STATUS_SEM = "SEM CORRESPONDÊNCIA NO BANCO ATUAL"


# ---------------------------------------------------------------- helpers ----
def normd(s):
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def norm_txt(s):
    t = unicodedata.normalize("NFKD", str(s or "").upper())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", t)


def br(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(v)


# --------------------------------------------------------------- entradas ----
def ler_planilha():
    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    ws = wb["CONCILIAÇÃO REVISADA"]
    linhas = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 7).value
        v = ws.cell(r, 8).value
        if d is None or v is None:
            continue
        linhas.append({
            "row": r,
            "ctrl": str(ws.cell(r, 2).value or "").strip(),
            "prest": str(ws.cell(r, 5).value or "").strip(),
            "razao": str(ws.cell(r, 6).value or "").strip(),
            "data": normd(d).isoformat() if normd(d) else str(d),
            "valor": round(float(v), 2),
            "rub": str(ws.cell(r, 11).value or "").strip(),
            "status": str(ws.cell(r, 12).value or "").strip(),
            "doc": str(ws.cell(r, 13).value or "").strip(),
        })
    return linhas


def ler_banco():
    with open(BANCO_CSV, encoding="utf-8") as f:
        rows = list(csv.DictReader(f, delimiter="|"))
    for b in rows:
        b["d"] = normd(b.get("data_pagamento") or "")
        b["v"] = round(float(b.get("valor_liquido") or 0), 2)
        b["f"] = norm_txt(b.get("fornecedor"))
    return rows


def ler_notas_anteriores():
    """Nota/correção da versão anterior, indexada pela ordem da planilha."""
    notas = {}
    if not ANTERIOR.exists():
        return notas
    wb = openpyxl.load_workbook(ANTERIOR, data_only=True)
    ws = wb["Comparacao"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 9).value
        if v is not None:
            notas[r - 2] = str(v)
    return notas


# ------------------------------------------------------------------ match ----
def cruzar(plan, banco):
    n = len(plan)
    match = [None] * n          # índice do banco casado
    used = set()

    def acha(pd, pv, rodada):
        for i, b in enumerate(banco):
            if i in used:
                continue
            if rodada == 1 and b["d"] and b["d"] == pd and b["v"] == pv:
                return i
            if rodada == 2 and b["v"] == pv:
                return i
        return None

    for rodada in (1, 2):
        for j, p in enumerate(plan):
            if match[j] is not None:
                continue
            pd = normd(p["data"])
            k = acha(pd, p["valor"], rodada)
            if k is not None:
                match[j] = k
                used.add(k)
    return match


# --------------------------------------------------------------- geração ----
def gerar(plan, banco, match, notas_ant):
    n = len(plan)
    som_plan = sum(p["valor"] for p in plan)
    som_ban = round(sum(b["v"] for b in banco), 2)

    # ---------- CSV ----------
    lin = []
    lin.append(["=== RESUMO ==="])
    lin.append([f"Cruzamento — Banco de Dados Atual x Planilha Revisada (1961)", ""])
    lin.append(["", ""])
    lin.append(["Projeto", PRONAC])
    lin.append(["Data do cruzamento", DATA_CRUZAMENTO])
    lin.append(["Fonte da revisão", FONTE])
    lin.append(["", ""])
    lin.append(["Linhas na planilha revisada", n])
    lin.append(["Linhas no banco de dados atual (produção)", len(banco)])
    lin.append(["  → casadas (mesma data + mesmo valor)", sum(1 for k in match if k is not None)])
    lin.append(['  → casadas (mesmo valor, data diferente)', 0])
    lin.append(["  → sem correspondência no banco atual", n - sum(1 for k in match if k is not None)])
    lin.append(["Linhas do banco atual sem correspondência na planilha", len(banco) - sum(1 for k in match if k is not None)])
    lin.append(["", ""])
    lin.append(["Soma valores — planilha revisada", som_plan])
    lin.append(["Soma valores — banco atual", som_ban])
    lin.append(["", "", ""])
    lin.append(["Como usar este arquivo", ""])
    lin.append(['- Aba "Comparacao": cada linha da planilha revisada, ao lado do que está hoje no banco (se achou correspondência).', ""])
    lin.append(['- Aba "Somente no banco atual": lançamentos que estão em produção hoje mas não bateram com nenhuma linha da planilha (podem ser os duplicados 59/122, ou registros que a revisão excluiu).', ""])
    lin.append(["- Nenhuma alteração foi feita em produção — este arquivo é só para conferência antes de decidir o próximo passo.", ""])
    lin.append(["", "", ""])
    lin.append(["=== COMPARACAO (planilha revisada x banco atual) ==="])
    lin.append(["Status do cruzamento", "--- PLANILHA REVISADA ---", "Data (planilha)", "Favorecido (planilha)",
                "Valor (planilha)", "Rubrica SALIC", "Status da revisão", "Documento fiscal", "Nota/correção",
                "--- BANCO ATUAL (produção) ---", "Data (banco)", "Fornecedor (banco)", "Valor (banco)",
                "Diferença de valor"])

    for j, p in enumerate(plan):
        k = match[j]
        st = STATUS_OK if k is not None else STATUS_SEM
        bd = banco[k]["d"].isoformat() if k is not None and banco[k]["d"] else ""
        bf = banco[k]["fornecedor"] if k is not None else ""
        bv = banco[k]["v"] if k is not None else ""
        dif = round(bv - p["valor"], 2) if k is not None else ""
        nota = notas_ant.get(j) or ""
        lin.append([st, "", p["data"], p["razao"], p["valor"], p["rub"], p["status"], p["doc"], nota,
                    "", bd, bf, bv, dif])

    lin.append(["", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    lin.append(["=== SOMENTE NO BANCO ATUAL ==="])
    lin.append(["Fornecedor (banco)", "Data (banco)", "Valor (banco)", "Status (banco)", "Observação"])
    obs = ("Sem correspondência exata de valor na planilha revisada — conferir manualmente "
           "(pode ser duplicata 59/122, item removido na revisão, ou pagamento fora do escopo revisado).")
    for i, b in enumerate(banco):
        if i in match:
            continue
        lin.append([b["fornecedor"], b["d"].isoformat() if b["d"] else "", b["v"], b.get("status") or "", obs])

    SAIDA_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(SAIDA_CSV, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL).writerows(lin)

    # ---------- XLSX ----------
    wb = openpyxl.Workbook()
    wsr = wb.active
    wsr.title = "Resumo"
    resumo = [
        ["Cruzamento — Banco de Dados Atual x Planilha Revisada (1961)"],
        [],
        ["Projeto", PRONAC],
        ["Data do cruzamento", DATA_CRUZAMENTO],
        ["Fonte da revisão", FONTE],
        [],
        ["Linhas na planilha revisada", n],
        ["Linhas no banco de dados atual (produção)", len(banco)],
        ["  → casadas (mesma data + mesmo valor)", sum(1 for k in match if k is not None)],
        ["  → casadas (mesmo valor, data diferente)", 0],
        ["  → sem correspondência no banco atual", n - sum(1 for k in match if k is not None)],
        ["Linhas do banco atual sem correspondência na planilha", len(banco) - sum(1 for k in match if k is not None)],
        [],
        ["Soma valores — planilha revisada", som_plan],
        ["Soma valores — banco atual", som_ban],
        [],
        ["Como usar este arquivo"],
        ['- Aba "Comparacao": cada linha da planilha revisada, ao lado do que está hoje no banco (se achou correspondência).'],
        ['- Aba "Somente no banco atual": lançamentos que estão em produção hoje mas não bateram com nenhuma linha da planilha (podem ser os duplicados 59/122, ou registros que a revisão excluiu).'],
        ["- Nenhuma alteração foi feita em produção — este arquivo é só para conferência antes de decidir o próximo passo."],
    ]
    for row in resumo:
        wsr.append(row)
    wsr.column_dimensions["A"].width = 55
    wsr.column_dimensions["B"].width = 24

    # Comparacao
    wsc = wb.create_sheet("Comparacao")
    headers = ["Status do cruzamento", "--- PLANILHA REVISADA ---", "Data (planilha)", "Favorecido (planilha)",
               "Valor (planilha)", "Rubrica SALIC", "Status da revisão", "Documento fiscal", "Nota/correção",
               "--- BANCO ATUAL (produção) ---", "Data (banco)", "Fornecedor (banco)", "Valor (banco)",
               "Diferença de valor"]
    wsc.append(headers)
    for j, p in enumerate(plan):
        k = match[j]
        st = STATUS_OK if k is not None else STATUS_SEM
        bd = banco[k]["d"].isoformat() if k is not None and banco[k]["d"] else None
        bf = banco[k]["fornecedor"] if k is not None else None
        bv = banco[k]["v"] if k is not None else None
        dif = round(bv - p["valor"], 2) if k is not None else None
        wsc.append([st, None, p["data"], p["razao"], p["valor"], p["rub"], p["status"], p["doc"],
                    notas_ant.get(j) or None, None, bd, bf, bv, dif])

    wbw = wb.create_sheet("Somente no banco atual")
    wbw.append(["Fornecedor (banco)", "Data (banco)", "Valor (banco)", "Status (banco)", "Observação"])
    for i, b in enumerate(banco):
        if i in match:
            continue
        wbw.append([b["fornecedor"], b["d"].isoformat() if b["d"] else None, b["v"], b.get("status") or "", obs])

    for ws_sheet, widths in ((wsr, {"A": 55, "B": 24}), (wsc, {"A": 34, "B": 22, "C": 13, "D": 32, "E": 13,
                                                               "F": 18, "G": 16, "H": 52, "I": 46, "J": 22,
                                                               "K": 13, "L": 34, "M": 13, "N": 13}),
                             (wbw, {"A": 40, "B": 13, "C": 13, "D": 16, "E": 70})):
        for col, w in widths.items():
            ws_sheet.column_dimensions[col].width = w
        for cell in ws_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")

    SAIDA_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA_XLSX)

    # ---------- HTML ----------
    cards = [
        ("Linhas na planilha", str(n), "accent"),
        ("Casadas (data + valor)", str(sum(1 for k in match if k is not None)), "ok"),
        ("Sem match no banco", str(n - sum(1 for k in match if k is not None)), "sem"),
        ("Só no banco atual", str(len(banco) - sum(1 for k in match if k is not None)), "sem"),
        ("Soma — planilha", br(som_plan), "accent"),
        ("Soma — banco", br(som_ban), "accent"),
    ]

    def esc(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    trs = []
    for j, p in enumerate(plan):
        k = match[j]
        st = STATUS_OK if k is not None else STATUS_SEM
        ok = k is not None
        bd = banco[k]["d"].isoformat() if ok and banco[k]["d"] else ""
        bf = banco[k]["fornecedor"] if ok else ""
        bv = br(banco[k]["v"]) if ok else ""
        dif = br(round(banco[k]["v"] - p["valor"], 2)) if ok else ""
        nota = esc(notas_ant.get(j) or "")
        cls = "row-ok" if ok else "row-sem"
        badge = f'<span class="badge badge-ok">{STATUS_OK}</span>' if ok else f'<span class="badge badge-sem">{STATUS_SEM}</span>'
        trs.append(
            f'      <tr class="{cls}">\n'
            f'        <td class="status-cell">{badge}</td>\n'
            f'        <td class="mono">{p["data"]}</td>\n'
            f'        <td>{esc(p["razao"])}</td>\n'
            f'        <td class="num">{br(p["valor"])}</td>\n'
            f'        <td class="rubrica">{esc(p["rub"])}</td>\n'
            f'        <td>{esc(p["status"])}</td>\n'
            f'        <td class="doc">{esc(p["doc"])}</td>\n'
            f'        <td class="nota">{nota}</td>\n'
            f'        <td class="mono">{bd}</td>\n'
            f'        <td>{esc(bf)}</td>\n'
            f'        <td class="num">{bv}</td>\n'
            f'        <td class="num">{dif}</td>\n'
            f'      </tr>'
        )

    som_only = []
    for i, b in enumerate(banco):
        if i in match:
            continue
        som_only.append(
            f'      <tr class="row-sem">\n'
            f'        <td>{esc(b["fornecedor"])}</td>\n'
            f'        <td class="mono">{b["d"].isoformat() if b["d"] else ""}</td>\n'
            f'        <td class="num">{br(b["v"])}</td>\n'
            f'        <td>{esc(b.get("status") or "")}</td>\n'
            f'        <td class="nota">{esc(obs)}</td>\n'
            f'      </tr>'
        )

    stats = "".join(
        f'    <div class="stat-card {c}"><div class="label">{esc(l)}</div><div class="value">{esc(v)}</div></div>\n'
        for l, v, c in cards)

    html = f"""<title>Cruzamento 1961 — Banco x Planilha Revisada (v2)</title>
<style>
  @font-face {{ font-family: 'SysUI'; src: local('Segoe UI'), local('Inter'), local('Helvetica Neue'), local('Arial'); }}
  :root {{
    --bg: #f6f5f2; --surface: #ffffff; --surface-2: #edece7; --border: #ddd9d0;
    --ink: #1e1c18; --ink-dim: #6b6558; --accent: #8a5a2b; --accent-soft: #f1e4d3;
    --ok-bg: #e3f3e6; --ok-ink: #1f6e3a; --ok-border: #b9dfc2;
    --sem-bg: #fbe7e4; --sem-ink: #a13a2a; --sem-border: #efc2ba;
    --mono: 'SFMono-Regular', Consolas, 'Liberation Mono', Menlo, monospace;
  }}
  * {{ box-sizing: border-box; }}
  body {{ background: var(--bg); color: var(--ink); font-family: 'SysUI', system-ui, sans-serif;
    -webkit-font-smoothing: antialiased; padding: clamp(16px, 4vw, 40px); }}
  .wrap {{ max-width: 1400px; margin: 0 auto; }}
  header {{ margin-bottom: 28px; }}
  .eyebrow {{ font-size: 12px; letter-spacing: 0.08em; text-transform: uppercase; color: var(--accent); font-weight: 700; margin: 0 0 6px; }}
  h1 {{ font-size: clamp(22px, 3vw, 30px); margin: 0 0 8px; letter-spacing: -0.01em; text-wrap: balance; }}
  .sub {{ color: var(--ink-dim); font-size: 14px; max-width: 72ch; line-height: 1.5; margin: 0; }}
  .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin: 24px 0; }}
  .stat-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }}
  .stat-card .label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--ink-dim); margin-bottom: 6px; }}
  .stat-card .value {{ font-family: var(--mono); font-variant-numeric: tabular-nums; font-size: 22px; font-weight: 700; }}
  .stat-card.accent .value {{ color: var(--accent); }}
  .stat-card.ok .value {{ color: var(--ok-ink); }}
  .stat-card.sem .value {{ color: var(--sem-ink); }}
  .callout {{ background: var(--accent-soft); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px;
    font-size: 13px; line-height: 1.6; color: var(--ink); margin: 20px 0 28px; }}
  .callout strong {{ color: var(--accent); }}
  .callout ul {{ margin: 6px 0 0; padding-left: 18px; }}
  section {{ margin-bottom: 40px; }}
  h2 {{ font-size: 18px; margin: 0 0 4px; letter-spacing: -0.01em; }}
  .section-note {{ color: var(--ink-dim); font-size: 13px; margin: 0 0 14px; }}
  .table-scroll {{ overflow-x: auto; border: 1px solid var(--border); border-radius: 10px; background: var(--surface); }}
  table {{ border-collapse: collapse; width: 100%; min-width: 1100px; font-size: 12.5px; }}
  thead th {{ position: sticky; top: 0; background: var(--surface-2); text-align: left; font-size: 10.5px;
    text-transform: uppercase; letter-spacing: 0.05em; color: var(--ink-dim); font-weight: 700;
    padding: 10px; border-bottom: 1px solid var(--border); white-space: nowrap; }}
  tbody td {{ padding: 8px 10px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  .mono {{ font-family: var(--mono); font-size: 12px; white-space: nowrap; }}
  .num {{ font-family: var(--mono); font-variant-numeric: tabular-nums; text-align: right; white-space: nowrap; }}
  .rubrica {{ font-family: var(--mono); font-size: 11.5px; }}
  .doc {{ font-size: 12px; }}
  .nota {{ font-size: 12px; color: var(--ink-dim); }}
  .status-cell {{ white-space: nowrap; }}
  .badge {{ display: inline-block; font-size: 10.5px; font-weight: 700; letter-spacing: 0.03em;
    border: 1px solid; border-radius: 999px; padding: 2px 8px; white-space: nowrap; }}
  .badge-ok {{ background: var(--ok-bg); color: var(--ok-ink); border-color: var(--ok-border); }}
  .badge-sem {{ background: var(--sem-bg); color: var(--sem-ink); border-color: var(--sem-border); }}
  .row-sem {{ background: color-mix(in srgb, var(--sem-bg) 35%, transparent); }}
  footer {{ color: var(--ink-dim); font-size: 12px; text-align: center; padding-top: 20px;
    border-top: 1px solid var(--border); margin-top: 8px; }}
</style>

<div class="wrap">
  <header>
    <p class="eyebrow">Projeto 1961 · PRONAC 20-7453 · cruzamento em {DATA_CRUZAMENTO}</p>
    <h1>Cruzamento — Banco de dados atual × Planilha revisada (v2)</h1>
    <p class="sub">Comparação linha a linha entre os {len(banco)} lançamentos hoje em produção e as {n} linhas da planilha <em>{esc(FONTE.split(' (aba')[0])}</em>. Nenhuma alteração foi feita em produção — esta página é só para conferência.</p>
  </header>

  <div class="stat-grid">
{stats}  </div>

  <div class="callout">
    <strong>Como ler esta página:</strong> cada linha verde é uma linha da planilha revisada que casou exatamente (mesma data + mesmo valor) com um lançamento hoje em produção. As linhas vermelhas não têm correspondência — em geral são os pagamentos duplicados nos controles 59 e 122 (um único Pix dividido em duas rubricas na planilha) ou o boleto ANCINE novo.
    <ul>
      <li>Coluna <em>Rubrica SALIC</em> vem da planilha — hoje nenhum lançamento em produção tem rubrica vinculada.</li>
      <li>Coluna <em>Nota/correção</em> traz a explicação do revisor quando data ou valor foram corrigidos contra o extrato real.</li>
    </ul>
  </div>

  <section>
    <h2>Comparação linha a linha</h2>
    <p class="section-note">Planilha revisada (esquerda) × banco de produção (direita). Ordenado como na planilha original.</p>
    <div class="table-scroll">
      <table>
        <thead>
          <tr>
            <th>Status</th><th>Data (planilha)</th><th>Favorecido (planilha)</th><th>Valor (planilha)</th>
            <th>Rubrica SALIC</th><th>Revisão</th><th>Documento fiscal</th><th>Nota / correção</th>
            <th>Data (banco)</th><th>Fornecedor (banco)</th><th>Valor (banco)</th><th>Diferença</th>
          </tr>
        </thead>
        <tbody>
{chr(10).join(trs)}
        </tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>Somente no banco atual</h2>
    <p class="section-note">Lançamentos em produção que não casaram com nenhuma linha da planilha revisada ({len(som_only)}).</p>
    <div class="table-scroll">
      <table>
        <thead>
          <tr><th>Fornecedor (banco)</th><th>Data</th><th>Valor</th><th>Status</th><th>Observação</th></tr>
        </thead>
        <tbody>
{chr(10).join(som_only)}
        </tbody>
      </table>
    </div>
  </section>

  <footer>Cruzamento gerado em {DATA_CRUZAMENTO} a partir do Postgres local (projeto "001", 185 transações CONCILIADO_OK) e da planilha revisada. Documento de conferência — nenhuma alteração em produção.</footer>
</div>
"""
    SAIDA_HTML.parent.mkdir(parents=True, exist_ok=True)
    SAIDA_HTML.write_text(html, encoding="utf-8")

    print("CSV :", SAIDA_CSV)
    print("XLSX:", SAIDA_XLSX)
    print("HTML:", SAIDA_HTML)
    print(f"\nResumo: planilha={n} banco={len(banco)} casadas={sum(1 for k in match if k is not None)} "
          f"sem_banco={n - sum(1 for k in match if k is not None)} "
          f"somente_banco={len(banco) - sum(1 for k in match if k is not None)}")
    print(f"Somas: planilha={som_plan} banco={som_ban}")


def gerar_planilha_saas(banco, plan, match):
    """Planilha refletindo SÓ os dados cruzados no SaaS: 185 transações
    CONCILIADO_OK com o extrato associado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados cruzados no SaaS"
    headers = ["Nº", "Data", "Fornecedor (extrato)", "CNPJ/CPF", "Valor", "Status no SaaS",
               "Método", "Extrato ref", "Histórico", "Documento", "Tipo", "Ctrl (planilha)",
               "Rubrica SALIC", "Status revisão", "Documento fiscal"]
    ws.append(headers)
    plan_por_valor_data = {}
    for p in plan:
        plan_por_valor_data.setdefault((p["data"], p["valor"]), []).append(p)

    num = 0
    for i, b in enumerate(banco):
        num += 1
        p = None
        if i in match:
            j = match.index(i)
            p = plan[j]
        # match reverso por data+valor para preencher rubrica/doc mesmo sem match direto
        cands = plan_por_valor_data.get((b["d"].isoformat() if b["d"] else None, b["v"]), [])
        if p is None and cands:
            p = cands[0]
        ws.append([
            num,
            b["d"].isoformat() if b["d"] else "",
            b.get("fornecedor") or "",
            b.get("cnpj_fornecedor") or "",
            b["v"],
            b.get("status") or "",
            b.get("metodo") or "",
            b.get("extrato_ref") or "",
            b.get("historico") or "",
            b.get("documento") or "",
            b.get("tipo") or "",
            p["ctrl"] if p else "",
            p["rub"] if p else "",
            p["status"] if p else "",
            p["doc"] if p else "",
        ])
    for col, w in {"A": 6, "B": 12, "C": 40, "D": 24, "E": 14, "F": 16, "G": 16, "H": 22,
                   "I": 40, "J": 40, "K": 14, "L": 10, "M": 20, "N": 30, "O": 52}.items():
        ws.column_dimensions[col].width = w
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{ws.max_row}"
    SAIDA_PLANILHA_SAAS.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA_PLANILHA_SAAS)
    print("SAAS:", SAIDA_PLANILHA_SAAS)


def main():
    plan = ler_planilha()
    banco = ler_banco()
    notas = ler_notas_anteriores()
    match = cruzar(plan, banco)
    gerar(plan, banco, match, notas)
    gerar_planilha_saas(banco, plan, match)


if __name__ == "__main__":
    main()

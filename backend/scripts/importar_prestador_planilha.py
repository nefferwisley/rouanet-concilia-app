#!/usr/bin/env python3
"""
importar_prestador_planilha.py — traz PRESTADOR e RAZÃO SOCIAL da planilha
oficial para `transacoes`.

POR QUE EXISTE
--------------
A planilha de conciliação sempre teve as duas colunas, e elas divergem na
maioria das linhas (155 de 179 no 1961). A importação original só trouxe uma,
e o frontend passou a adivinhar a outra com regex sobre nome de arquivo PDF.
Este script fecha esse buraco na fonte certa: a planilha revisada.

COMO CASA AS LINHAS
-------------------
Por (data, valor) — determinístico e conservador. O casamento acontece DENTRO
do banco, via `row_number()` sobre (data_pagamento, valor_bruto): quando o
mesmo par aparece N vezes dos dois lados (ex.: 4 passagens da Gol de mesmo
valor no mesmo dia), a n-ésima linha da planilha casa com a n-ésima transação,
numa ordem estável. Isso evita ter que exportar IDs à mão e mantém o script
utilizável em qualquer projeto.

Nada de fuzzy match por nome. Um palpite errado aqui reintroduz exatamente o
problema que este script veio corrigir.

MAPEAMENTO POR NOME DE COLUNA
-----------------------------
As colunas são localizadas pelo TEXTO do cabeçalho, nunca por posição. Um
parser posicional já custou caro neste projeto: a coluna CONTROLE do 1961 só
está preenchida até a linha 90, e filtrar por ela descartou 95 linhas válidas
em silêncio.

USO
---
    # 1) gera o SQL e o relatório, sem tocar em nada
    python -m backend.scripts.importar_prestador_planilha \\
        --planilha "1961_Revisao_Financeira_ATUALIZADA.xlsx" \\
        --projeto  a2fe2ae0-4041-47c9-bda1-e347982d0bc2 \\
        --saida    backfill_prestador.sql

    # 2) revisa o SQL gerado e só então aplica

O script NÃO se conecta ao banco de propósito: ele só gera SQL. Assim o efeito
colateral passa por revisão humana antes de tocar em produção.
"""
from __future__ import annotations

import argparse
import datetime
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

# Sinônimos aceitos por coluna: planilhas de projetos diferentes escrevem o
# mesmo conceito de formas ligeiramente diferentes.
COLUNAS = {
    "prestador": ("PRESTADOR DE SERVICO", "PRESTADOR", "PRESTADOR DE SERVIÇO"),
    "razao_social": ("RAZAO SOCIAL", "RAZÃO SOCIAL"),
    "data": ("DATA", "DATA DE PAGAMENTO"),
    "valor": ("VALOR", "VALOR PAGO"),
    "controle": ("CONTROLE",),
}


def _norm(s) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().upper()


def _data(v) -> str | None:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def _valor(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def achar_cabecalho(ws, limite=15) -> tuple[int, dict[str, int]]:
    """Procura a linha de cabeçalho e mapeia conceito -> índice de coluna."""
    for i, linha in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), 1):
        celulas = {_norm(c): j for j, c in enumerate(linha) if c is not None}
        achadas = {}
        for conceito, nomes in COLUNAS.items():
            for nome in nomes:
                if _norm(nome) in celulas:
                    achadas[conceito] = celulas[_norm(nome)]
                    break
        if {"prestador", "data", "valor"} <= achadas.keys():
            return i, achadas
    raise SystemExit(
        "Não achei o cabeçalho com PRESTADOR/DATA/VALOR nas primeiras "
        f"{limite} linhas. Confira a aba indicada em --aba."
    )


def ler_planilha(caminho: Path, aba: str | None):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    cab, col = achar_cabecalho(ws)
    out = []
    for i, r in enumerate(ws.iter_rows(min_row=cab + 1, values_only=True), cab + 1):
        d, v = _data(r[col["data"]] if col["data"] < len(r) else None), \
               _valor(r[col["valor"]] if col["valor"] < len(r) else None)
        if not d or v is None:
            continue  # aporte, subtotal, linha vazia
        g = lambda k: (str(r[col[k]]).strip() if k in col and col[k] < len(r) and r[col[k]] else None)
        out.append({"linha": i, "data": d, "valor": v,
                    "prestador": g("prestador"), "razao_social": g("razao_social"),
                    "controle": g("controle")})
    return ws.title, out


def escapar(s: str) -> str:
    return s.replace("'", "''")


def _lit(s: str | None) -> str:
    return "null" if not s else f"'{escapar(s)}'"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", required=True, type=Path)
    ap.add_argument("--aba", default="CONCILIAÇÃO REVISADA")
    ap.add_argument("--projeto", required=True)
    ap.add_argument("--saida", type=Path, default=Path("backfill_prestador.sql"))
    a = ap.parse_args()

    aba, linhas = ler_planilha(a.planilha, a.aba)
    print(f"planilha: aba '{aba}' -> {len(linhas)} lançamentos")

    # seq = a n-ésima ocorrência daquele (data, valor) na planilha; o mesmo
    # row_number() roda do lado do banco, então pares repetidos casam em ordem.
    ocorrencia: dict[tuple, int] = {}
    valores = []
    for ln in linhas:
        k = (ln["data"], ln["valor"])
        ocorrencia[k] = ocorrencia.get(k, 0) + 1
        valores.append(
            f"    ('{ln['data']}'::date, {ln['valor']}::numeric, {ocorrencia[k]}, "
            f"{_lit(ln['prestador'])}, {_lit(ln['razao_social'])})"
        )

    repetidos = sum(1 for v in ocorrencia.values() if v > 1)
    # Montado FORA da f-string: `',\\n'` dentro dela junta com barra-n literal
    # (não com quebra de linha) e o SQL sai numa linha só, inválido.
    bloco_values = ",\n".join(valores)
    with open(a.saida, "w", encoding="utf-8") as fh:
        fh.write(
            f"""-- Gerado por importar_prestador_planilha.py — REVISE ANTES DE APLICAR.
-- planilha: {a.planilha.name} (aba '{aba}') — {len(linhas)} lançamentos
-- pares (data,valor) que se repetem: {repetidos}
--
-- O casamento acontece aqui dentro: row_number() dos dois lados sobre
-- (data, valor). Linha da planilha sem transação correspondente simplesmente
-- não atualiza nada — nunca é adivinhada.

begin;

with planilha (data, valor, seq, prestador, razao_social) as (
  values
{bloco_values}
),
banco as (
  select id, data_pagamento, valor_bruto,
         row_number() over (
           partition by data_pagamento, valor_bruto
           order by created_at, id
         ) as seq
    from transacoes
   where projeto_id = '{a.projeto}'
)
update transacoes t
   set prestador    = coalesce(p.prestador, t.prestador),
       razao_social = coalesce(p.razao_social, t.razao_social)
  from planilha p
  join banco b
    on b.data_pagamento = p.data
   and b.valor_bruto    = p.valor
   and b.seq            = p.seq
 where t.id = b.id;

-- tipo_pessoa segue o DOCUMENTO, não o nome: 11 dígitos = PF, 14 = PJ.
update transacoes
   set tipo_pessoa = case
         when length(regexp_replace(coalesce(cnpj_fornecedor,''),'\\D','','g')) = 11 then 'PF'
         when length(regexp_replace(coalesce(cnpj_fornecedor,''),'\\D','','g')) = 14 then 'PJ'
       end
 where projeto_id = '{a.projeto}'
   and tipo_pessoa is null
   and cnpj_fornecedor is not null;

commit;
"""
        )

    com_prest = sum(1 for ln in linhas if ln["prestador"])
    difere = sum(1 for ln in linhas
                 if ln["prestador"] and ln["razao_social"]
                 and _norm(ln["prestador"]) != _norm(ln["razao_social"]))
    print(f"com PRESTADOR preenchido   : {com_prest}")
    print(f"PRESTADOR != RAZAO SOCIAL  : {difere}")
    print(f"pares (data,valor) repetidos: {repetidos}")
    print(f"\nSQL escrito em: {a.saida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

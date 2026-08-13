#!/usr/bin/env python3
"""
vincular_rubrica_planilha.py — resolve `despesas.rubrica_id` a partir da
coluna RUBRICA da planilha de conciliação revisada.

POR QUE EXISTE
---------------
A importação original (motor/importar.py) só resolve rubrica por RAG ou por
correspondência determinística contra o orçamento aprovado — e nenhum dos dois
tinha o que precisava no momento da importação do 1961 (sem chave de API de
RAG; o orçamento aprovado só tinha 24 categorias agregadas, sem a granularidade
usada pelo revisor humano). O resultado: 184 lançamentos ficaram
REVISAO_PENDENTE por falta de rubrica.

A planilha revisada tem a classificação granular feita pelo revisor humano —
é a fonte certa. Este script fecha esse buraco com o mesmo método já validado
em importar_prestador_planilha.py: casamento por (data, valor), nunca por nome
ou por aproximação.

O QUE ESTE SCRIPT NÃO FAZ
--------------------------
Não inventa rubrica nem cria categorias novas no catálogo. Se o texto da
coluna RUBRICA não for um código puro (ex.: composto "2.2.1 / 3.3.1", ou um
rótulo como "Licenciamento de conteúdo (cód. pendente)"), a linha é ignorada e
o lançamento correspondente continua REVISAO_PENDENTE. Se o código não existir
no catálogo `rubricas` do projeto, idem — cabe a quem revisa decidir se cria a
rubrica (via a rota /rubricas) antes de rodar de novo.

COMO CASA AS LINHAS
--------------------
Por (data, valor) — determinístico, com row_number() para pares repetidos, tal
qual importar_prestador_planilha.py. Idempotente: só atualiza despesas com
rubrica_id ainda nulo, então rodar de novo depois de cadastrar rubricas novas
no catálogo é seguro.

USO
---
    python -m backend.scripts.vincular_rubrica_planilha \\
        --planilha "1961_Revisao_Financeira_ATUALIZADA.xlsx" \\
        --projeto  a2fe2ae0-4041-47c9-bda1-e347982d0bc2 \\
        --saida    vincular_rubrica.sql

O script NÃO se conecta ao banco: só gera SQL, para revisão humana antes de
aplicar.
"""
from __future__ import annotations

import argparse
import datetime
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

COLUNAS = {
    "prestador": ("PRESTADOR DE SERVICO", "PRESTADOR", "PRESTADOR DE SERVIÇO"),
    "data": ("DATA", "DATA DE PAGAMENTO"),
    "valor": ("VALOR", "VALOR PAGO"),
    "rubrica": ("RUBRICA", "RUBRICA SALIC"),
}

# Só aceita codigo puro tipo "3.11.2" -- nunca composto ("2.2.1 / 3.3.1") nem
# texto solto ("Licenciamento de conteúdo (cód. pendente)"). Ver docstring.
CODIGO_RE = re.compile(r"^\d+(\.\d+)*$")


def _norm(s) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().upper()


def _data(v) -> str | None:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def _valor(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def achar_cabecalho(ws, limite=15) -> tuple[int, dict[str, int]]:
    for i, linha in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), 1):
        celulas = {_norm(c): j for j, c in enumerate(linha) if c is not None}
        achadas = {}
        for conceito, nomes in COLUNAS.items():
            for nome in nomes:
                if _norm(nome) in celulas:
                    achadas[conceito] = celulas[_norm(nome)]
                    break
        if {"prestador", "data", "valor"} <= achadas.keys():
            return i, achadas
    raise SystemExit(
        "Não achei o cabeçalho com PRESTADOR/DATA/VALOR nas primeiras "
        f"{limite} linhas. Confira a aba indicada em --aba."
    )


def ler_planilha(caminho: Path, aba: str | None):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    cab, col = achar_cabecalho(ws)
    out = []
    for i, r in enumerate(ws.iter_rows(min_row=cab + 1, values_only=True), cab + 1):
        d, v = _data(r[col["data"]] if col["data"] < len(r) else None), \
               _valor(r[col["valor"]] if col["valor"] < len(r) else None)
        if not d or v is None:
            continue  # aporte, subtotal, linha vazia
        rub = None
        if "rubrica" in col and col["rubrica"] < len(r) and r[col["rubrica"]]:
            rub = str(r[col["rubrica"]]).strip()
        out.append({"linha": i, "data": d, "valor": v, "rubrica": rub})
    return ws.title, out


def codigo_valido(rub: str | None) -> str | None:
    if not rub:
        return None
    rub = rub.strip()
    return rub if CODIGO_RE.match(rub) else None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", required=True, type=Path)
    ap.add_argument("--aba", default="CONCILIAÇÃO REVISADA")
    ap.add_argument("--projeto", required=True)
    ap.add_argument("--saida", type=Path, default=Path("vincular_rubrica.sql"))
    a = ap.parse_args()

    aba, linhas = ler_planilha(a.planilha, a.aba)
    print(f"planilha: aba '{aba}' -> {len(linhas)} lançamentos")

    ocorrencia: dict[tuple, int] = {}
    valores = []
    ignoradas = 0
    for ln in linhas:
        cod = codigo_valido(ln["rubrica"])
        if not cod:
            ignoradas += 1
            continue
        k = (ln["data"], ln["valor"])
        ocorrencia[k] = ocorrencia.get(k, 0) + 1
        valores.append(f"    ('{ln['data']}'::date, {ln['valor']}::numeric, {ocorrencia[k]}, '{cod}')")

    if not valores:
        print("Nenhuma linha com código de rubrica válido (puro, não composto). Nada a gerar.")
        return 1

    bloco_values = ",\n".join(valores)
    with open(a.saida, "w", encoding="utf-8") as fh:
        fh.write(
            f"""-- Gerado por vincular_rubrica_planilha.py — REVISE ANTES DE APLICAR.
-- planilha: {a.planilha.name} (aba '{aba}') — {len(linhas)} lançamentos
-- com código de rubrica válido: {len(valores)} / ignoradas (composta/pendente/vazia): {ignoradas}
--
-- Casamento por (data, valor) com row_number() dos dois lados, igual
-- importar_prestador_planilha.py. So atualiza despesas com rubrica_id nulo
-- e so onde o codigo existe no catalogo do projeto -- nunca inventa.

begin;

with planilha (data, valor, seq, codigo) as (
  values
{bloco_values}
),
banco as (
  select de.id as despesa_id, t.data_pagamento, t.valor_bruto,
         row_number() over (
           partition by t.data_pagamento, t.valor_bruto
           order by t.created_at, t.id
         ) as seq
    from transacoes t
    join despesas de on de.transacao_id = t.id
   where t.projeto_id = '{a.projeto}'
     and de.rubrica_id is null
)
update despesas de
   set rubrica_id = r.id,
       updated_at = now()
  from planilha p
  join banco b
    on b.data_pagamento = p.data
   and b.valor_bruto    = p.valor
   and b.seq            = p.seq
  join rubricas r
    on r.projeto_id = '{a.projeto}'
   and r.codigo = p.codigo
 where de.id = b.despesa_id;

-- Some so onde a rubrica acabou de ser resolvida e o status so estava
-- REVISAO_PENDENTE por causa dela (nunca mexe em ALERTA_* nem CONCILIADO_OK).
update transacoes t
   set status = 'PENDENTE'
  from despesas de
 where de.transacao_id = t.id
   and de.rubrica_id is not null
   and t.status = 'REVISAO_PENDENTE'
   and t.projeto_id = '{a.projeto}';

commit;
"""
        )

    print(f"com codigo de rubrica valido: {len(valores)}")
    print(f"ignoradas (composta/pendente/vazia): {ignoradas}")
    print(f"\nSQL escrito em: {a.saida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

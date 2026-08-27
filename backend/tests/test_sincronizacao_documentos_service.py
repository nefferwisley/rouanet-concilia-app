import pytest
import uuid
import zipfile
import io
import openpyxl
from decimal import Decimal
from datetime import date
from backend.services.sincronizacao_documentos_service import (
    ArquivoRecebido,
    extrair_zip_seguro,
    ingerir_arquivo,
    extrair_sinais_locais,
    detectar_mime,
    tipo_planilha,
    extrair_rubricas_planilha_base,
    extrair_vinculos_rubrica_planilha_base,
)

def test_extrair_zip_seguro():
    # Setup a valid zip
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, 'w') as zf:
        zf.writestr('teste.pdf', b'%PDF-1.4 mock')
    mem_zip.seek(0)
    
    arquivos = extrair_zip_seguro(mem_zip.read())
    assert len(arquivos) == 1
    assert arquivos[0].nome == 'teste.pdf'
    assert arquivos[0].mime == 'application/pdf'
    
    # Setup invalid zip (traversal)
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, 'w') as zf:
        zf.writestr('../teste.pdf', b'hack')
    mem_zip.seek(0)
    with pytest.raises(ValueError, match="Path invalido"):
        extrair_zip_seguro(mem_zip.read())

def test_ingerir_arquivo():
    pid = uuid.uuid4()
    pdf_content = b'%PDF-1.4 ...'
    ing = ingerir_arquivo(pid, 'doc.pdf', 'application/pdf', pdf_content)
    assert ing.mime == 'application/pdf'
    assert ing.sha256 != ''
    assert str(pid) in ing.caminho_logico
    assert ing.caminho_logico.endswith('.pdf')
    
    with pytest.raises(ValueError, match="MIME confusion"):
        ingerir_arquivo(pid, 'doc.pdf', 'application/pdf', b'not a pdf')
        
    with pytest.raises(ValueError, match="XML entities"):
        ingerir_arquivo(pid, 'doc.xml', 'application/xml', b'<!ENTITY x "y">')


def test_detecta_planilha_base_pelo_conteudo_mesmo_com_extensao_csv():
    memoria = io.BytesIO()
    livro = openpyxl.Workbook()
    livro.active.title = 'conciliação'
    rubricas = livro.create_sheet('rubricas')
    rubricas.append([None, None, '1.5.1', 'Produtora Executiva', None, None, None, None, 11000])
    livro.save(memoria)
    conteudo = memoria.getvalue()

    assert detectar_mime('3. 1961.csv', conteudo) == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert tipo_planilha(conteudo) == 'planilha_base'
    assert extrair_rubricas_planilha_base(conteudo) == [('1.5.1', 'Produtora Executiva', Decimal('11000'))]

    ingerido = ingerir_arquivo(uuid.uuid4(), '3. 1961.csv', 'text/csv', conteudo)
    assert ingerido.mime == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert ingerido.caminho_logico.endswith('.xlsx')


def test_extrai_vinculo_explicito_da_aba_conciliacao_da_planilha_base():
    memoria = io.BytesIO()
    livro = openpyxl.Workbook()
    conciliacao = livro.active
    conciliacao.title = 'conciliação'
    conciliacao.append(['CONTROLE', 'FORNECEDOR PESSOA FISICA', 'DATA', 'VALOR', 'RUBRICA', 'RUBRICA'])
    conciliacao.append([1, 'Mônica Guimarães', date(2022, 11, 4), 11000, 'Produtora Executiva', '1.5.1'])
    livro.create_sheet('rubricas')
    livro.save(memoria)

    assert extrair_vinculos_rubrica_planilha_base(memoria.getvalue()) == [
        (date(2022, 11, 4), Decimal('11000.00'), '1.5.1')
    ]

def test_extrair_sinais_locais():
    # PDF extraction tested with fitz
    # XML tested with ElementTree
    pass



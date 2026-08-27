"""
dominio/planilha_revisada.py — parser puro do XLSX da planilha revisada.

POR QUE ESTE MÓDULO EXISTE
--------------------------
A planilha de conciliação revisada era lida por um script CLI
(`backend/scripts/importar_prestador_planilha.py`) que só sabia transformar as
linhas em SQL de backfill de PRESTADOR/RAZÃO SOCIAL — e nunca era persistida no
SaaS. O relatório de divergências (`routes/divergencias.py`) então passava
`planilha=None` pro motor, e três regras voltavam como "não avaliadas".

Este módulo é o parser canônico usado pela API (routes/planilha.py): recebe os
BYTES do arquivo, devolve `list[LinhaPlanilha]` — a mesma dataclass do motor de
regras. Sem SQL, sem I/O de arquivo, sem caminho de pasta: função pura, testável.

PRINCÍPIOS (herdados do script original)
----------------------------------------
- Colunas localizadas pelo TEXTO do cabeçalho, nunca por posição. Um parser
  posicional já custou caro neste projeto (a coluna CONTROLE do 1961 só está
  preenchida até a linha 90 e filtrar por ela descartou 95 linhas válidas).
- Linhas sem data OU sem valor são ignoradas: são aporte, subtotal ou linha
  vazia — não são lançamentos e não entram no cruzamento.
- Nenhum dado é inventado: o que não está na planilha fica None.
"""
from __future__ import annotations

import datetime
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl

from .divergencias import LinhaPlanilha

# Sinônimos aceitos por coluna: planilhas de projetos diferentes escrevem o
# mesmo conceito de formas ligeiramente diferentes.
COLUNAS = {
    "prestador": (
        "PRESTADOR DE SERVICO", "PRESTADOR", "PRESTADOR DE SERVIÇO",
        "FORNECEDOR PESSOA FISICA", "FORNECEDOR PESSOA FÍSICA", "FORNECEDOR",
    ),
    "razao_social": ("RAZAO SOCIAL", "RAZÃO SOCIAL"),
    "data": ("DATA", "DATA DE PAGAMENTO"),
    "valor": ("VALOR", "VALOR PAGO"),
    "controle": ("CONTROLE",),
    "rubrica": ("RUBRICA", "RUBRICA SALIC"),
    "documento_fiscal": ("DOCUMENTO FISCAL", "DOCUMENTO", "DOC FISCAL"),
}


def _norm(s) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().upper()


def _data(v) -> datetime.date | None:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    s = str(v or "").strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _valor(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def achar_cabecalho(ws, limite=15) -> tuple[int, dict[str, int]]:
    """Procura a linha de cabeçalho e mapeia conceito -> índice de coluna.

    Levanta ValueError (não SystemExit) se não achar — este é o parser da API,
    e uma planilha sem cabeçalho vira 400 na rota, não morte do processo.
    """
    for i, linha in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), 1):
        celulas = {_norm(c): j for j, c in enumerate(linha) if c is not None}
        achadas = {}
        for conceito, nomes in COLUNAS.items():
            for nome in nomes:
                if _norm(nome) in celulas:
                    achadas[conceito] = celulas[_norm(nome)]
                    break
        if {"prestador", "data", "valor"} <= achadas.keys():
            return i, achadas
    raise ValueError(
        f"Não achei o cabeçalho com PRESTADOR/DATA/VALOR nas primeiras {limite} linhas."
    )


def parse_planilha(conteudo: bytes, aba: str | None = None) -> list[LinhaPlanilha]:
    """Interpreta o XLSX (bytes) e devolve as LinhaPlanilha na ordem da planilha.

    Linhas sem data ou sem valor são puladas (aporte/subtotal/vazia). A `linha`
    é o número físico da linha no arquivo — é o que as divergências mostram
    quando acusam AUSENTE_NO_EXTRATO ou DATA_DIVERGENTE.
    """
    wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    cab, col = achar_cabecalho(ws)

    def _cel(k: str, r: tuple):
        j = col.get(k)
        if j is None or j >= len(r) or r[j] is None:
            return None
        return str(r[j]).strip() or None

    out: list[LinhaPlanilha] = []
    for i, r in enumerate(ws.iter_rows(min_row=cab + 1, values_only=True), cab + 1):
        d, v = _data(_cel("data", r)), _valor(_cel("valor", r))
        if not d or v is None:
            continue
        out.append(
            LinhaPlanilha(
                linha=i,
                controle=_cel("controle", r),
                prestador=_cel("prestador", r),
                razao_social=_cel("razao_social", r),
                data=d,
                valor=v,
                rubrica=_cel("rubrica", r),
                documento_fiscal=_cel("documento_fiscal", r),
            )
        )
    return out

import asyncio
import hashlib
import io
import uuid
import zipfile
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal

import asyncpg

from backend.services.storage_service import criar_arquivo_se_ausente, remover_arquivo
from backend.services.storage_service import logger as storage_logger
import fitz
import openpyxl
from backend.dominio.planilha_revisada import parse_planilha

@dataclass
class SinaisDocumento:
    texto_extraido: str | None = None
    erro_extracao: str | None = None
    valor: Decimal | None = None
    data_documento: date | None = None
    cpf_cnpj: str | None = None
    numero_documento: str | None = None
    favorecido_normalizado: str | None = None
    tipo_documental: str | None = None

@dataclass
class ArquivoRecebido:
    nome: str
    mime: str
    conteudo: bytes

@dataclass
class ArquivoIngerido:
    nome: str
    mime: str
    sha256: str
    tamanho_bytes: int
    conteudo: bytes
    caminho_logico: str


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _eh_xlsx(conteudo: bytes) -> bool:
    """Reconhece Office Open XML pelo conteúdo, não pelo sufixo do arquivo."""
    if not zipfile.is_zipfile(io.BytesIO(conteudo)):
        return False
    with zipfile.ZipFile(io.BytesIO(conteudo)) as arquivo:
        nomes = set(arquivo.namelist())
    return "[Content_Types].xml" in nomes and "xl/workbook.xml" in nomes


def detectar_mime(nome: str, conteudo: bytes, mime_informado: str | None = None) -> str:
    """Preserva planilhas-base mesmo quando chegaram com extensão incorreta."""
    if _eh_xlsx(conteudo):
        return MIME_XLSX
    ext = nome.rsplit(".", 1)[-1].lower() if "." in nome else ""
    if ext == "pdf":
        return "application/pdf"
    if ext == "xml":
        return "application/xml"
    if ext in ("png", "jpg", "jpeg"):
        return f"image/{ext.replace('jpg', 'jpeg')}"
    return mime_informado or "application/octet-stream"


def tipo_planilha(conteudo: bytes) -> str:
    """Classifica a planilha-base pelo conteúdo da aba, sem confiar no nome."""
    if not _eh_xlsx(conteudo):
        return "planilha"
    livro = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    try:
        abas = {
            "".join(c for c in unicodedata.normalize("NFKD", aba.lower()) if not unicodedata.combining(c))
            for aba in livro.sheetnames
        }
    finally:
        livro.close()
    return "planilha_base" if "rubricas" in abas else "planilha"


def extrair_rubricas_planilha_base(conteudo: bytes) -> list[tuple[str, str, Decimal | None]]:
    """Lê as rubricas de uma planilha-base como a do Projeto 1961.

    Apenas códigos de item são importados; linhas agregadoras (etapas) ficam
    fora do catálogo porque não recebem pagamentos diretamente.
    """
    livro = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    try:
        aba = next((nome for nome in livro.sheetnames if unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("ascii").lower() == "rubricas"), None)
        if aba is None:
            return []
        rubricas = []
        for linha in livro[aba].iter_rows(min_row=1, values_only=True):
            codigo = str(linha[2] or "").strip() if len(linha) > 2 else ""
            descricao = str(linha[3] or "").strip() if len(linha) > 3 else ""
            if not re.fullmatch(r"\d+(?:\.\d+){2,}", codigo) or not descricao:
                continue
            valor = linha[8] if len(linha) > 8 else None
            try:
                valor = Decimal(str(valor)) if valor is not None else None
            except Exception:
                valor = None
            rubricas.append((codigo, descricao, valor))
        return rubricas
    finally:
        livro.close()


def extrair_vinculos_rubrica_planilha_base(conteudo: bytes) -> list[tuple[date, Decimal, str]]:
    """Obtém somente vínculos explícitos por data, valor e código de rubrica."""
    vinculos = []
    for linha in parse_planilha(conteudo):
        codigo = (linha.rubrica or "").strip()
        if re.fullmatch(r"\d+(?:\.\d+)*", codigo):
            vinculos.append((linha.data, linha.valor, codigo))
    return vinculos


async def vincular_despesas_planilha_base(
    conn: asyncpg.Connection,
    projeto_id: uuid.UUID,
    vinculos: list[tuple[date, Decimal, str]],
) -> int:
    """Vincula apenas despesas ainda sem rubrica, preservando as já classificadas."""
    if not vinculos:
        return 0
    atualizadas = await conn.fetch(
        """
        with planilha_bruta as (
          select data, valor, codigo, ordem
            from unnest($2::date[], $3::numeric[], $4::text[]) with ordinality
                 as p(data, valor, codigo, ordem)
        ), planilha as (
          select data, valor, codigo,
                 row_number() over (partition by data, valor order by ordem) as sequencia
            from planilha_bruta
        ), banco as (
          select d.id as despesa_id, t.id as transacao_id, t.data_pagamento, t.valor_bruto,
                 row_number() over (
                   partition by t.data_pagamento, t.valor_bruto order by t.created_at, t.id
                 ) as sequencia
            from transacoes t
            join despesas d on d.transacao_id = t.id
           where t.projeto_id = $1 and d.rubrica_id is null
        )
        update despesas d
           set rubrica_id = r.id, updated_at = now()
          from planilha p
          join banco b on b.data_pagamento = p.data
                      and b.valor_bruto = p.valor
                      and b.sequencia = p.sequencia
          join rubricas r on r.projeto_id = $1 and r.codigo = p.codigo
         where d.id = b.despesa_id
        returning b.transacao_id
        """,
        projeto_id,
        [v[0] for v in vinculos],
        [v[1] for v in vinculos],
        [v[2] for v in vinculos],
    )
    for linha in atualizadas:
        await conn.execute(
            """
            update transacoes set status = 'PENDENTE'
             where id = $1 and projeto_id = $2 and status = 'REVISAO_PENDENTE'
            """,
            linha["transacao_id"], projeto_id,
        )
    return len(atualizadas)

async def iniciar_sincronizacao(conn: asyncpg.Connection, projeto_id: uuid.UUID, user_id: uuid.UUID, arquivos: list[ArquivoRecebido]) -> uuid.UUID:
    row = await conn.fetchrow(
        '''
        INSERT INTO sincronizacoes_documentos (projeto_id, criado_por, status, recebidos)
        VALUES ($1, $2, 'recebendo', $3)
        RETURNING id
        ''',
        projeto_id, user_id, len(arquivos)
    )
    sinc_id = row['id']
    return sinc_id


def expandir_arquivos_recebidos(arquivos: list[ArquivoRecebido]) -> list[ArquivoRecebido]:
    """Expande ZIPs de pasta, mas nunca confunde uma planilha XLSX com um ZIP."""
    resultado: list[ArquivoRecebido] = []
    for arquivo in arquivos:
        mime = detectar_mime(arquivo.nome, arquivo.conteudo, arquivo.mime)
        if mime != MIME_XLSX and zipfile.is_zipfile(io.BytesIO(arquivo.conteudo)):
            resultado.extend(extrair_zip_seguro(arquivo.conteudo))
        else:
            resultado.append(ArquivoRecebido(arquivo.nome, mime, arquivo.conteudo))
    return resultado


async def registrar_arquivos_sincronizacao(
    conn: asyncpg.Connection,
    sincronizacao_id: uuid.UUID,
    projeto_id: uuid.UUID,
    arquivos: list[ArquivoRecebido],
) -> None:
    """Persiste todos os arquivos recebidos antes do processamento em segundo plano."""
    arquivos = expandir_arquivos_recebidos(arquivos)
    deduplicados = 0
    for arquivo in arquivos:
        ingerido = ingerir_arquivo(projeto_id, arquivo.nome, arquivo.mime, arquivo.conteudo)
        storage_key, _ = criar_arquivo_se_ausente(ingerido.caminho_logico, ingerido.conteudo)
        documento_id = await conn.fetchval(
            """
            insert into documentos_sincronizacao
              (sincronizacao_id, projeto_id, sha256, storage_key, nome_exibicao,
               mime_type, tamanho_bytes)
            values ($1, $2, $3, $4, $5, $6, $7)
            on conflict (projeto_id, sha256) do nothing
            returning id
            """,
            sincronizacao_id, projeto_id, ingerido.sha256, storage_key, ingerido.nome,
            ingerido.mime, ingerido.tamanho_bytes,
        )
        if documento_id is None:
            deduplicados += 1
    await conn.execute(
        """
        update sincronizacoes_documentos
           set recebidos = $2, deduplicados = $3
         where id = $1
        """,
        sincronizacao_id, len(arquivos), deduplicados,
    )

def extrair_zip_seguro(conteudo: bytes, max_uncompressed_bytes: int = 250 * 1024 * 1024) -> list[ArquivoRecebido]:
    arquivos = []
    total_uncompressed = 0
    with io.BytesIO(conteudo) as b:
        if not zipfile.is_zipfile(b):
            raise ValueError("Nao e um arquivo ZIP valido")
        b.seek(0)
        with zipfile.ZipFile(b, 'r') as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                # Block symlink
                if info.create_system == 3 and (info.external_attr >> 16) & 0o120000 == 0o120000:
                    raise ValueError("Symlinks nao permitidos")
                
                nome = info.filename
                if '\0' in nome or '..' in nome or nome.startswith('/') or nome.startswith('\\'):
                    raise ValueError(f"Path invalido no ZIP: {nome}")
                
                total_uncompressed += info.file_size
                if total_uncompressed > max_uncompressed_bytes:
                    raise ValueError("ZIP excede o limite de expansao")
                
                file_content = zf.read(info.filename)
                
                mime = detectar_mime(nome, file_content)
                arquivos.append(ArquivoRecebido(nome=nome, mime=mime, conteudo=file_content))
    return arquivos

def ingerir_arquivo(projeto_id: uuid.UUID, nome: str, mime: str, conteudo: bytes) -> ArquivoIngerido:
    if not conteudo:
        raise ValueError("Arquivo vazio")
    
    tamanho = len(conteudo)
    sha256 = hashlib.sha256(conteudo).hexdigest()
    
    mime = detectar_mime(nome, conteudo, mime)
    extensao = 'bin'
    if mime == 'application/pdf' or nome.lower().endswith('.pdf'):
        if not conteudo.startswith(b'%PDF-'):
            raise ValueError("MIME confusion: nuo e um PDF valido")
        extensao = 'pdf'
        mime = 'application/pdf'
    elif mime in ('application/xml', 'text/xml') or nome.lower().endswith('.xml'):
        if b'<!ENTITY' in conteudo:
            raise ValueError("XML entities not allowed")
        extensao = 'xml'
        mime = 'application/xml'
    elif mime.startswith('image/'):
        extensao = mime.split('/')[-1]
    elif mime == MIME_XLSX:
        extensao = 'xlsx'
    
    # Chave imutavel
    caminho = f"{projeto_id}/sincronizacao/{sha256}.{extensao}"
    
    return ArquivoIngerido(
        nome=nome,
        mime=mime,
        sha256=sha256,
        tamanho_bytes=tamanho,
        conteudo=conteudo,
        caminho_logico=caminho
    )
def extrair_sinais_locais(nome: str, mime: str, conteudo: bytes) -> SinaisDocumento:
    texto = ""
    erro = None
    
    try:
        if mime == 'application/pdf':
            doc = fitz.open(stream=conteudo, filetype="pdf")
            paginas = []
            for i in range(min(5, doc.page_count)):  # Limite para nuo estourar memoria
                paginas.append(doc[i].get_text())
            texto = "\n".join(paginas)
            doc.close()
        elif mime == 'application/xml':
            import xml.etree.ElementTree as ET
            # XML seguro validado no ingestor
            root = ET.fromstring(conteudo.decode('utf-8', errors='ignore'))
            # Extrair texto de tags comuns de NF
            textos = []
            for elem in root.iter():
                if elem.text and elem.text.strip():
                    textos.append(elem.text.strip())
            texto = " ".join(textos)
    except Exception as e:
        erro = f"Falha na extracao local: {str(e)}"
    
    # Extrair sinais regex simples (CPF/CNPJ, valor, data)
    sinais = SinaisDocumento(texto_extraido=texto, erro_extracao=erro)
    if mime == MIME_XLSX:
        sinais.tipo_documental = tipo_planilha(conteudo)
    
    # Busca heuristica bosica no texto ou no nome
    alvo = (texto + " " + nome).replace('\n', ' ')
    
    # Valor (R$ 1.234,56 ou 1234.56)
    m_valor = re.search(r'(?:R\$|Valor)\s*[:]?\s*(\d{1,3}(?:\.\d{3})*,\d{2}|\d+\.\d{2})', alvo, re.IGNORECASE)
    if m_valor:
        val_str = m_valor.group(1).replace('.', '').replace(',', '.')
        try:
            sinais.valor = Decimal(val_str)
        except:
            pass
            
    # CPF/CNPJ
    m_doc = re.search(r'\b(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{3}\.\d{3}\.\d{3}-\d{2}|\d{11}|\d{14})\b', alvo)
    if m_doc:
        sinais.cpf_cnpj = re.sub(r'\D', '', m_doc.group(1))
        
    return sinais

from backend.database import adquirir_conn

async def processar_sincronizacao(sincronizacao_id: uuid.UUID) -> None:
    pool, conn_bg = await adquirir_conn()
    if conn_bg is None:
        return
        
    transacao = None
    commit_iniciado = False
    lock_acquired = False
    
    try:
        # Get projeto_id
        row = await conn_bg.fetchrow(
            'SELECT projeto_id FROM sincronizacoes_documentos WHERE id = $1',
            sincronizacao_id
        )
        if not row:
            return
            
        projeto_id = row['projeto_id']
        lock_key = int.from_bytes(hashlib.sha256(str(projeto_id).encode()).digest()[:8], "little", signed=True)
        
        # O lock advisory e transacional
        transacao = conn_bg.transaction()
        await transacao.start()
        
        lock_acquired = await conn_bg.fetchval(
            "SELECT pg_try_advisory_xact_lock($1)", lock_key
        )
        if not lock_acquired:
            await conn_bg.execute(
                "UPDATE sincronizacoes_documentos SET status = 'erro', erro_operacional = 'Lock ocupado' WHERE id = $1",
                sincronizacao_id
            )
            await transacao.rollback()
            return
            
        await conn_bg.execute(
            "UPDATE sincronizacoes_documentos SET status = 'processando' WHERE id = $1",
            sincronizacao_id
        )
        
        # 2. Extracao local (produz sinais)
        from backend.services.storage_service import baixar_arquivo
        
        # Get pending documents for this sync
        docs = await conn_bg.fetch(
            '''
            SELECT id, storage_key, nome_exibicao, mime_type
            FROM documentos_sincronizacao
            WHERE sincronizacao_id = $1 AND estado_extracao = 'pendente'
            ''',
            sincronizacao_id
        )
        
        for d in docs:
            try:
                conteudo = baixar_arquivo(d['storage_key'])
                if conteudo is None:
                    raise ValueError("Arquivo não encontrado no storage")
                sinais = extrair_sinais_locais(d['nome_exibicao'], d['mime_type'], conteudo)
                if sinais.tipo_documental == 'planilha_base':
                    for codigo, descricao, valor_orcado in extrair_rubricas_planilha_base(conteudo):
                        await conn_bg.execute(
                            """
                            insert into rubricas (projeto_id, codigo, descricao, valor_orcado)
                            values ($1, $2, $3, $4)
                            on conflict (projeto_id, codigo) do update
                              set descricao = excluded.descricao,
                                  valor_orcado = excluded.valor_orcado
                            """,
                            projeto_id, codigo, descricao, valor_orcado,
                        )
                    await vincular_despesas_planilha_base(
                        conn_bg,
                        projeto_id,
                        extrair_vinculos_rubrica_planilha_base(conteudo),
                    )
                
                await conn_bg.execute(
                    '''
                    UPDATE documentos_sincronizacao
                    SET estado_extracao = 'extraido',
                        erro_extracao = $2,
                        tipo_documental = $3,
                        valor = $4,
                        data_documento = $5,
                        cpf_cnpj = $6,
                        favorecido_normalizado = $7
                    WHERE id = $1
                    ''',
                    d['id'],
                    sinais.erro_extracao,
                    sinais.tipo_documental or 'desconhecido',
                    sinais.valor,
                    sinais.data_documento,
                    sinais.cpf_cnpj,
                    sinais.favorecido_normalizado
                )
            except Exception as e:
                await conn_bg.execute(
                    '''
                    UPDATE documentos_sincronizacao
                    SET estado_extracao = 'erro', erro_extracao = $2
                    WHERE id = $1
                    ''',
                    d['id'], str(e)
                )
                
        # 3. Geracao deterministica de candidatos (Matching)
        # TODO: call matching logic here
        
        await conn_bg.execute(
            "UPDATE sincronizacoes_documentos SET status = 'concluida' WHERE id = $1",
            sincronizacao_id
        )
        
        commit_iniciado = True
        await transacao.commit()
    except Exception as e:
        if transacao is not None and not commit_iniciado:
            await transacao.rollback()
        # TODO: Ambiguous commit handling (like W2-T5)
        # But wait, starting the sync already uploaded the file!
        # If the file upload was in iniciar_sincronizacao and it already committed,
        # there are NO files uploaded inside this transaction!
        # Thus, we don't have new orphans to compensate in processar_sincronizacao.
        pass
    finally:
        await pool.release(conn_bg)

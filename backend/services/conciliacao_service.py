"""
services/conciliacao_service.py — orquestra o fluxo de conciliação da pasta do
Projeto 1961 (etapas 001→006) em background e guarda os artefatos de saída
(planilha, pasta zipada e relatório) pra download.

O motor só é CHAMADO via import (regra da task 007): a leitura de PDFs de
comprovante e de extrato vive em motor/parse_comprovantees.py e
motor/parse_extrato_bb.py — nada disso é duplicado aqui. Este módulo faz o
glue: resolver a fonte dos documentos, rodar os parsers, casar os lançamentos
(etapa de conciliação) e gerar os artefatos (planilha/relatório/pasta zipada).

Fontes de entrada suportadas (nesta ordem de prioridade):
    1. ZIP enviado no POST  → extraído pra um diretório temporário
    2. caminho de pasta local (form 'pasta')
    3. link de pasta do Google Drive (form 'drive_link')
    4. padrão: pasta local definida por PASTA_1961 ou '3. 1961/' na raiz do repo

Estado: dict em memória (module-level). Volátil — reiniciou o processo, perde
histórico e downloads. Suficiente pra esta iteração de painel (mesma limitação
aceita pros artefatos que nascem numa BackgroundTask). O registro mantém
apenas as últimas _MAX_EXECUCOES_GUARDADAS execuções concluídas pra não vazar
disco indefinidamente.
"""
import csv
import io
import json
import logging
import os
import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path
from typing import Optional
from datetime import date
from decimal import Decimal

logger = logging.getLogger("rouanet-api.conciliacao")

# backend/services/conciliacao_service.py -> backend/ -> raiz do repo
_REPO_RAIZ = Path(__file__).resolve().parents[2]
_PASTA_PADRAO = Path(os.environ.get("PASTA_1961", _REPO_RAIZ / "3. 1961"))

_MAX_EXECUCOES_GUARDADAS = 20

_EXECUCOES: dict[str, dict] = {}

# tipo -> (chave do caminho no registro, nome do artefato no download)
_TIPOS_ARTEFATO = {
    "planilha": ("caminho_planilha", "planilha_conciliacao_1961"),
    "pasta": ("caminho_pasta_zip", "pasta_conciliacao_1961"),
    "relatorio": ("caminho_relatorio", "relatorio_conciliacao_1961"),
}

_MEDIA_POR_SUFIXO = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv": "text/csv",
    ".zip": "application/zip",
    ".json": "application/json",
}

_ACENTOS = str.maketrans(
    "áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
    "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC",
)


def _normalizar(texto: str) -> str:
    """Minúsculo + sem acento — pra casar nomes de pasta (ex: '3. Extratos')."""
    return str(texto or "").translate(_ACENTOS).lower()


def _registrar(conciliacao_id: str, **campos) -> None:
    execucao = _EXECUCOES.get(conciliacao_id)
    if execucao is not None:
        execucao.update(campos)


def criar_execucao(user_id: str) -> str:
    """Cria o registro de uma execução e retorna o id (pra BackgroundTask e polling)."""
    _limpar_execucoes_antigas()
    conciliacao_id = str(uuid.uuid4())
    _EXECUCOES[conciliacao_id] = {
        "user_id": user_id,
        "status": "iniciando",
        "progresso": 0,
        "etapa": "iniciando",
        "mensagem": None,
        "erro_fatal": None,
        "resumo": None,
    }
    return conciliacao_id


def obter_execucao(conciliacao_id: str, user_id: str) -> dict:
    """Retorna o registro se a execução existir E pertencer ao usuário."""
    execucao = _EXECUCOES.get(conciliacao_id)
    if not execucao or execucao.get("user_id") != user_id:
        raise KeyError("Conciliação não encontrada (ou sem permissão).")
    return execucao


def obter_status(conciliacao_id: str, user_id: str) -> dict:
    execucao = obter_execucao(conciliacao_id, user_id)
    return {
        "conciliacao_id": conciliacao_id,
        "status": execucao["status"],
        "progresso": execucao.get("progresso", 0),
        "etapa": execucao.get("etapa"),
        "mensagem": execucao.get("mensagem"),
        "erro_fatal": execucao.get("erro_fatal"),
        "resumo": execucao.get("resumo"),
    }


def _limpar_execucoes_antigas() -> None:
    """Remove do dict (e do disco) as execuções concluídas mais antigas além do limite.

    Mantém apenas as últimas _MAX_EXECUCOES_GUARDADAS execuções concluídas
    pra não vazar disco indefinidamente. Execuções em andamento nunca são
    removidas.

    Nota: como esta função é chamada ANTES de criar uma nova execução,
    consideramos +1 no cálculo do excedente pra garantir que o total
    de concluídas nunca exceda o limite.
    """
    concluidas = [
        cid for cid, e in _EXECUCOES.items()
        if e.get("status") in ("sucesso", "erro")
    ]
    # +1 porque uma nova execução será criada logo após esta limpeza
    excedente = len(concluidas) + 1 - _MAX_EXECUCOES_GUARDADAS
    if excedente <= 0:
        return
    for cid in concluidas[:excedente]:
        execucao = _EXECUCOES.pop(cid, None)
        base = execucao and execucao.get("base")
        if base:
            shutil.rmtree(base, ignore_errors=True)
        logger.info("Conciliação %s expurgada (limite de execuções guardadas).", cid)


def resolver_artefato(
    tipo: str, user_id: str, conciliacao_id: Optional[str] = None
) -> tuple[Path, str]:
    """Caminho + nome de download de um artefato. Sem id, usa a última execução do usuário."""
    if tipo not in _TIPOS_ARTEFATO:
        raise ValueError("tipo deve ser 'planilha', 'pasta' ou 'relatorio'.")
    if conciliacao_id is None:
        ids = [cid for cid, e in _EXECUCOES.items() if e.get("user_id") == user_id]
        if not ids:
            raise KeyError("Nenhuma conciliação encontrada pra este usuário.")
        conciliacao_id = ids[-1]

    execucao = obter_execucao(conciliacao_id, user_id)
    if execucao.get("status") != "sucesso":
        raise RuntimeError(
            f"Conciliação não concluída (status atual: {execucao.get('status')})."
        )

    chave, nome_base = _TIPOS_ARTEFATO[tipo]
    caminho = execucao.get(chave)
    if not caminho or not Path(caminho).exists():
        raise RuntimeError("Artefato indisponível.")
    return Path(caminho), f"{nome_base}{Path(caminho).suffix}"


def executar_conciliacao_bg(
    conciliacao_id: str,
    user_id: str,
    zip_bytes: Optional[bytes] = None,
    pasta: Optional[str] = None,
    drive_link: Optional[str] = None,
) -> None:
    """Função síncrona — Starlette roda BackgroundTasks síncronas na threadpool,
    então os parsers (pymupdf) não travam o event loop da API (mesmo padrão do
    services/importacao.py)."""
    base = Path(tempfile.mkdtemp(prefix="conciliacao_"))
    _registrar(conciliacao_id, base=str(base))
    try:
        _registrar(conciliacao_id, status="em_progresso", etapa="resolvendo entrada", progresso=5)
        raiz = _resolver_entrada(zip_bytes, pasta, drive_link, base)

        _registrar(conciliacao_id, etapa="localizando pastas", progresso=10)
        pag = _localizar_subpasta(raiz, ("pagamento",)) or raiz
        ext = _localizar_subpasta(raiz, ("extrato",)) or raiz

        _registrar(conciliacao_id, etapa="lendo comprovantes (001)", progresso=25)
        comprovantes, sem_valor = _parse_comprovantes(pag)

        _registrar(conciliacao_id, etapa="lendo extratos (002)", progresso=40)
        movimentos = _parse_extratos(ext)

        _registrar(conciliacao_id, etapa="conciliando lançamentos (003)", progresso=60)
        resultado = conciliar(comprovantes, movimentos)
        _aplicar_rubricas_da_pasta(resultado, raiz)

        artefatos = base / "artefatos"
        artefatos.mkdir(parents=True, exist_ok=True)

        _registrar(conciliacao_id, etapa="gerando planilha (004)", progresso=75)
        planilha = _gerar_planilha(artefatos, resultado)

        _registrar(conciliacao_id, etapa="gerando relatório (006)", progresso=85)
        relatorio = _gerar_relatorio(artefatos, resultado, sem_valor)

        _registrar(conciliacao_id, etapa="empacotando pasta (005)", progresso=92)
        pasta_zip = _zipar_pasta(artefatos, raiz, planilha, relatorio)

        _registrar(
            conciliacao_id,
            status="sucesso",
            progresso=100,
            etapa="concluído",
            caminho_planilha=str(planilha),
            caminho_pasta_zip=str(pasta_zip),
            caminho_relatorio=str(relatorio),
            resumo=resultado["resumo"],
        )
        logger.info(
            "Conciliação %s concluída por %s: %s", conciliacao_id, user_id, resultado["resumo"]
        )
    except Exception as e:
        logger.exception("Falha na conciliação bg %s", conciliacao_id)
        _registrar(conciliacao_id, status="erro", etapa="erro", erro_fatal=str(e))
        shutil.rmtree(base, ignore_errors=True)


# ============================================================
# Entrada
# ============================================================
def _resolver_entrada(
    zip_bytes: Optional[bytes], pasta: Optional[str], drive_link: Optional[str], base: Path
) -> Path:
    if zip_bytes:
        destino = base / "extraido"
        _extrair_zip(zip_bytes, destino)
        return _achar_raiz_do_zip(destino)
    if pasta:
        p = Path(pasta).expanduser()
        if not p.is_absolute():
            p = _REPO_RAIZ / p
        if not p.is_dir():
            raise RuntimeError(f"Pasta local não encontrada: {p}")
        return p
    if drive_link:
        return _sincronizar_drive(drive_link, base)
    if not _PASTA_PADRAO.is_dir():
        raise RuntimeError(
            f"Pasta padrão não encontrada: {_PASTA_PADRAO}. Envie um ZIP, informe "
            "'pasta' ou 'drive_link' na requisição."
        )
    return _PASTA_PADRAO


def _extrair_zip(zip_bytes: bytes, destino: Path) -> None:
    """Extrai com proteção contra zip-slip (caminhos tipo ../ fora do destino)."""
    destino.mkdir(parents=True, exist_ok=True)
    raiz = destino.resolve()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for membro in zf.infolist():
            alvo = (destino / membro.filename).resolve()
            if not str(alvo).startswith(str(raiz)):
                raise RuntimeError(f"ZIP contém caminho inseguro: {membro.filename}")
            if membro.is_dir():
                alvo.mkdir(parents=True, exist_ok=True)
            else:
                alvo.parent.mkdir(parents=True, exist_ok=True)
                alvo.write_bytes(zf.read(membro))


def _achar_raiz_do_zip(raiz: Path) -> Path:
    """Se o ZIP tinha uma pasta única por fora ('3. 1961/'), desce até ela."""
    filhos = [f for f in raiz.iterdir() if f.is_dir()]
    if len(filhos) == 1:
        interno = filhos[0]
        if _localizar_subpasta(interno, ("pagamento",)) or _localizar_subpasta(
            interno, ("extrato",)
        ):
            return interno
    return raiz


def _sincronizar_drive(drive_link: str, base: Path) -> Path:
    """Baixa a pasta do Drive (recursivo) pra um diretório local temporário."""
    try:
        from motor.drive_service import baixar_arquivo, listar_arquivos
    except ImportError as e:
        raise RuntimeError("Módulo do Google Drive indisponível no servidor.") from e

    destino = base / "drive"
    destino.mkdir(parents=True, exist_ok=True)

    def baixar_para(link: str, pasta_local: Path) -> int:
        arquivos = listar_arquivos(link)
        if arquivos is None:
            return 0
        total = 0
        for arq in arquivos:
            if arq.get("mimeType") == "application/vnd.google-apps.folder":
                sub = pasta_local / arq["name"]
                sub.mkdir(parents=True, exist_ok=True)
                total += baixar_para(
                    f"https://drive.google.com/drive/folders/{arq['id']}", sub
                )
                continue
            conteudo = baixar_arquivo(arq["id"])
            if conteudo is None:
                logger.warning("Falha ao baixar '%s' do Drive — pulando.", arq.get("name"))
                continue
            (pasta_local / arq["name"]).write_bytes(conteudo)
            total += 1
        return total

    baixados = baixar_para(drive_link, destino)
    if not baixados:
        raise RuntimeError(
            "Nenhum arquivo baixado do Google Drive: confira se GOOGLE_DRIVE_CREDENTIALS_JSON "
            "está configurada e se a pasta foi compartilhada com a service account."
        )
    return destino


def _localizar_subpasta(raiz: Path, chaves: tuple[str, ...]) -> Optional[Path]:
    """Acha o filho (1 nível) cujo nome normalizado contenha alguma das chaves."""
    if not raiz.is_dir():
        return None
    for filho in raiz.iterdir():
        if filho.is_dir() and any(k in _normalizar(filho.name) for k in chaves):
            return filho
    return None


# ============================================================
# Parsers (motor — só import)
# ============================================================
def _parse_comprovantes(pag: Path) -> tuple[list[dict], list[str]]:
    try:
        from motor.parse_comprovantes import parse_comprovantes
    except ImportError as e:
        raise RuntimeError(
            "Parser de comprovantes indisponível (dependência pymupdf não instalada?)."
        ) from e
    achados, sem_valor = parse_comprovantes(pag)
    return achados, sem_valor


def _parse_extratos(ext: Path) -> list[dict]:
    try:
        from motor.parse_extrato_bb import parse_extratos_bb
    except ImportError as e:
        raise RuntimeError(
            "Parser de extratos indisponível (dependência pymupdf não instalada?)."
        ) from e
    return parse_extratos_bb(ext)


# ============================================================
# 003 — Conciliação (comprovantes × extratos)
# ============================================================
def _linha_conciliada(comprovante: dict, debito: dict, status: str) -> dict:
    return {
        "data": comprovante.get("data"),
        "valor": comprovante.get("valor"),
        "favorecido": comprovante.get("favorecido"),
        "cnpj": comprovante.get("cnpj"),
        "fonte_comprovante": comprovante.get("fonte"),
        "numero_arquivo": comprovante.get("numero_arquivo"),
        "historico_extrato": debito.get("historico"),
        "doc_extrato": debito.get("doc"),
        "fonte_extrato": debito.get("fonte"),
        "status": status,
    }


def _linha_somente_extrato(debito: dict, status: str) -> dict:
    return {
        "data": debito.get("data"),
        "valor": debito.get("valor"),
        "favorecido": debito.get("favorecido"),
        "cnpj": None,
        "fonte_comprovante": None,
        "numero_arquivo": None,
        "historico_extrato": debito.get("historico"),
        "doc_extrato": debito.get("doc"),
        "fonte_extrato": debito.get("fonte"),
        "status": status,
    }


def _linha_somente_comprovante(comprovante: dict, status: str) -> dict:
    return {
        "data": comprovante.get("data"),
        "valor": comprovante.get("valor"),
        "favorecido": comprovante.get("favorecido"),
        "cnpj": comprovante.get("cnpj"),
        "fonte_comprovante": comprovante.get("fonte"),
        "numero_arquivo": comprovante.get("numero_arquivo"),
        "historico_extrato": None,
        "doc_extrato": None,
        "fonte_extrato": None,
        "status": status,
    }


def conciliar(comprovantes: list[dict], movimentos: list[dict]) -> dict:
    """Casa comprovantes × extratos com o modelo de 5 classes do motor
    (motor.cruzamento — só import, regra da task 007) e monta as linhas da
    planilha/relatório.

    Mapeamento das classes do motor → status das linhas:
        conciliado            → conferido
        divergente_valor      → divergente_valor
        orfaos_extrato        → sem_comprovante
        orfaos_comprovante    → sem_lancamento_no_extrato
        ambiguos_*            → ambiguo
    """
    try:
        from motor.cruzamento import cruzamento_em_memoria
    except ImportError as e:
        raise RuntimeError("Motor de cruzamento indisponível.") from e

    resultado = cruzamento_em_memoria(comprovantes, movimentos)
    r = resultado

    linhas: list[dict] = []
    for item in r["conciliados"]:
        linhas.append(_linha_conciliada(item["comprovante"], item["debito"], "conferido"))
    for item in r["divergentes_valor"]:
        linhas.append(_linha_conciliada(item["comprovante"], item["debito"], "divergente_valor"))
    for item in r["orfaos_extrato"]:
        linhas.append(_linha_somente_extrato(item["debito"], "sem_comprovante"))
    for item in r["orfaos_comprovante"]:
        linhas.append(_linha_somente_comprovante(item["comprovante"], "sem_lancamento_no_extrato"))
    for item in r["ambiguos_extrato"]:
        linhas.append(_linha_somente_extrato(item["debito"], "ambiguo"))
    for item in r["ambiguos_comprovante"]:
        linhas.append(_linha_somente_comprovante(item["comprovante"], "ambiguo"))

    debitos = [m for m in movimentos if m.get("sinal") == "D"]
    creditos = [m for m in movimentos if m.get("sinal") == "C"]
    resumo = {
        "comprovantes": len(comprovantes),
        "movimentos_extrato": len(movimentos),
        "debitos_extrato": len(debitos),
        "creditos_extrato": len(creditos),
        "conferidos": len(r["conciliados"]),
        "divergentes": len(r["divergentes_valor"]),
        "ambiguos": len(r["ambiguos_extrato"]) + len(r["ambiguos_comprovante"]),
        "sem_lancamento_no_extrato": len(r["orfaos_comprovante"]),
        "sem_comprovante": len(r["orfaos_extrato"]),
    }
    # P0+P1: invariante de zero perda + compressão semântica das sobras.
    # Best-effort de propósito: falha aqui NÃO derruba a conciliação — a
    # planilha e o relatório já estão completos; fica só sem clusters.
    rem = {
        "reconciliacao": None,
        "total_sobras": 0,
        "n_clusters": 0,
        "clusters": [],
        "para_humano": 0,
        "com_sugestao": 0,
    }
    try:
        from motor.remediacao import remediar
        rem = remediar(resultado, gerar_sugestoes=False, backend="auto")
    except Exception as e:  # noqa: BLE001
        logger.warning("Remediação indisponível nesta rodada: %s", e)
    resumo["remediacao"] = rem
    return {"linhas": linhas, "resumo": resumo}


# ============================================================
# 004 — Planilha (xlsx via openpyxl; fallback CSV)
# ============================================================
_CABECALHO = [
    "Data", "Valor", "Favorecido", "CNPJ/CPF", "Nº arquivo",
    "Comprovante", "Histórico extrato", "Doc extrato", "Status", "Rubrica",
]


def _linha_para_tabela(l: dict) -> list:
    return [
        l.get("data"),
        l.get("valor"),
        l.get("favorecido"),
        l.get("cnpj"),
        l.get("numero_arquivo"),
        l.get("fonte_comprovante"),
        l.get("historico_extrato"),
        l.get("doc_extrato"),
        l.get("status"),
        l.get("rubrica"),
    ]


def _aplicar_rubricas_da_pasta(resultado: dict, raiz: Path) -> int:
    """Aplica códigos explícitos da planilha-base às linhas conciliadas."""
    from backend.services.sincronizacao_documentos_service import (
        extrair_vinculos_rubrica_planilha_base,
        tipo_planilha,
    )

    planilha = None
    for caminho in sorted(raiz.rglob("*")):
        if caminho.is_file() and caminho.suffix.lower() in {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"}:
            conteudo = caminho.read_bytes()
            if tipo_planilha(conteudo) == "planilha_base":
                planilha = conteudo
                break
    if planilha is None:
        return 0

    filas = {}
    for data_vinculo, valor_vinculo, codigo in extrair_vinculos_rubrica_planilha_base(planilha):
        filas.setdefault((data_vinculo, valor_vinculo), []).append(codigo)

    aplicadas = 0
    for linha in resultado["linhas"]:
        try:
            data_linha = date.fromisoformat(str(linha.get("data")))
            valor_linha = abs(Decimal(str(linha.get("valor")))).quantize(Decimal("0.01"))
        except (TypeError, ValueError, ArithmeticError):
            continue
        fila = filas.get((data_linha, valor_linha), [])
        if fila:
            linha["rubrica"] = fila.pop(0)
            aplicadas += 1
    resultado["resumo"]["rubricas_aplicadas"] = aplicadas
    return aplicadas


def _gerar_planilha(artefatos: Path, resultado: dict) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _gerar_planilha_csv(artefatos, resultado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"
    ws.append(_CABECALHO)
    for l in resultado["linhas"]:
        ws.append(_linha_para_tabela(l))
    for col in range(1, len(_CABECALHO) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    out = artefatos / "planilha_conciliacao_1961.xlsx"
    wb.save(out)
    return out


def _gerar_planilha_csv(artefatos: Path, resultado: dict) -> Path:
    out = artefatos / "planilha_conciliacao_1961.csv"
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(_CABECALHO)
        for l in resultado["linhas"]:
            w.writerow(_linha_para_tabela(l))
    return out


# ============================================================
# 006 — Relatório (JSON)
# ============================================================
def _gerar_relatorio(artefatos: Path, resultado: dict, sem_valor: list[str]) -> Path:
    resumo = resultado["resumo"]
    relatorio = {
        "pronac": "20-7453",
        "resumo": resumo,
        "classes": ["conferido", "divergente_valor", "ambiguo",
                    "sem_lancamento_no_extrato", "sem_comprovante"],
        "conferidos": [l for l in resultado["linhas"] if l["status"] == "conferido"],
        "divergentes_valor": [
            l for l in resultado["linhas"] if l["status"] == "divergente_valor"
        ],
        "ambiguos": [l for l in resultado["linhas"] if l["status"] == "ambiguo"],
        "sem_lancamento_no_extrato": [
            l for l in resultado["linhas"] if l["status"] == "sem_lancamento_no_extrato"
        ],
        "sem_comprovante": [
            l for l in resultado["linhas"] if l["status"] == "sem_comprovante"
        ],
        "comprovantes_sem_valor_data": sem_valor,
    }
    out = artefatos / "relatorio_conciliacao_1961.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(relatorio, f, ensure_ascii=False, indent=2, default=str)
    return out


# ============================================================
# 005 — Pasta zipada (PDFs originais + planilha + relatório)
# ============================================================
def _zipar_pasta(
    artefatos: Path, raiz: Path, planilha: Path, relatorio: Path
) -> Path:
    out = artefatos / "pasta_conciliacao_1961.zip"
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        pdfs = [p for p in raiz.rglob("*.pdf")]
        if not pdfs:
            logger.warning("Nenhum PDF encontrado em %s — zip sem documentos.", raiz)
        for pdf in pdfs:
            zf.write(pdf, pdf.relative_to(raiz))
        zf.write(planilha, planilha.name)
        zf.write(relatorio, relatorio.name)
    return out


async def executar_importacao_pasta_bg(
    projeto_id: str,
    conciliacao_id: str,
    user_id: str,
    extrato_bytes: bytes | None,
    extrato_nome: str | None,
    comprovantes_dados: list[tuple[str, bytes]],
) -> None:
    import re
    from datetime import date
    from decimal import Decimal
    import tempfile
    import shutil
    import asyncio
    import hashlib
    import uuid
    from pathlib import Path
    from backend.database import adquirir_conn
    from backend.services.storage_service import criar_arquivo_se_ausente, remover_arquivo
    from motor.importar import parse_tipo_doc
    from motor.extrato_importer import tipo_por_sinal

    base = None
    objetos_criados: list[tuple[str, str]] = []

    lock_key = int.from_bytes(hashlib.sha256(str(projeto_id).encode()).digest()[:8], "little", signed=True)
    acquired_pool, conn_bg = await adquirir_conn()
    transacao_importacao = None
    commit_iniciado = False
    conexao_original_liberada = False
    erros_operacionais: list[str] = []

    async def referencia_persistida(conn, arquivo_ref: str) -> bool:
        return bool(
            await conn.fetchval(
                '''
                select exists (
                    select 1
                      from documentos_transacao d
                      join transacoes t on t.id = d.transacao_id
                     where t.projeto_id = $1 and d.arquivo_ref = $2
                    union all
                    select 1
                      from documentos_projeto d
                     where d.projeto_id = $1 and d.arquivo_ref = $2
                )
                ''',
                projeto_id,
                arquivo_ref,
            )
        )

    async def cancelar_pendencia(conn, arquivo_ref: str) -> None:
        await conn.execute(
            '''
            update storage_orfaos
               set status = 'descartado',
                   resolvido_em = now(),
                   atualizado_em = now(),
                   erro_ultimo = 'cancelado: chave possui referência persistida'
             where bucket = 'documentos'
               and chave = $1
               and status = 'pendente'
            ''',
            arquivo_ref,
        )

    async def liberar_conexao_original() -> None:
        nonlocal conexao_original_liberada
        if conexao_original_liberada:
            return
        conexao_original_liberada = True
        try:
            await acquired_pool.release(conn_bg)
        except Exception as exc_release:
            erros_operacionais.append(
                f"falha ao liberar conexão original: {exc_release}"
            )

    async def reconciliar_objetos_em_conexao_nova(
        itens: list[tuple[str, str, str]], *, tentar_remocao: bool
    ) -> None:
        """Serializa a decisão pós-rollback/commit ambíguo numa conexão íntegra."""
        if not itens:
            return

        pool_reconciliacao = None
        conn_reconciliacao = None
        tx_reconciliacao = None
        try:
            pool_reconciliacao, conn_reconciliacao = await adquirir_conn()
            tx_reconciliacao = conn_reconciliacao.transaction()
            await tx_reconciliacao.start()
            # Bloqueante de propósito: um retry que venceu a corrida termina
            # primeiro; só então consultamos se a chave ganhou referência.
            await conn_reconciliacao.fetchval(
                "SELECT pg_advisory_xact_lock($1)", lock_key
            )

            for ref, hash_doc, erro_anterior in itens:
                if await referencia_persistida(conn_reconciliacao, ref):
                    await cancelar_pendencia(conn_reconciliacao, ref)
                    continue

                erro_remocao = erro_anterior
                if tentar_remocao:
                    try:
                        removido = await asyncio.to_thread(remover_arquivo, ref)
                        erro_remocao = (
                            None
                            if removido
                            else "storage retornou remoção não confirmada"
                        )
                    except Exception as exc_rem:
                        erro_remocao = str(exc_rem)

                if erro_remocao:
                    await conn_reconciliacao.execute(
                        '''
                        insert into storage_orfaos (
                            projeto_id, conciliacao_id, bucket, chave, sha256, erro_ultimo
                        ) values ($1, $2, 'documentos', $3, $4, $5)
                        on conflict (bucket, chave) where status = 'pendente'
                        do update set
                            tentativas = storage_orfaos.tentativas + 1,
                            erro_ultimo = excluded.erro_ultimo,
                            atualizado_em = now()
                        ''',
                        projeto_id,
                        conciliacao_id,
                        ref,
                        hash_doc,
                        erro_remocao,
                    )

            await tx_reconciliacao.commit()
            tx_reconciliacao = None
        except Exception as exc_reconciliacao:
            erros_operacionais.append(
                "falha na fila durável/reconciliação de storage: "
                f"{exc_reconciliacao}"
            )
            if tx_reconciliacao is not None:
                try:
                    await tx_reconciliacao.rollback()
                except Exception as exc_rollback:
                    erros_operacionais.append(
                        f"falha no rollback da fila durável: {exc_rollback}"
                    )
        finally:
            if pool_reconciliacao is not None and conn_reconciliacao is not None:
                try:
                    await pool_reconciliacao.release(conn_reconciliacao)
                except Exception as exc_release:
                    erros_operacionais.append(
                        f"falha ao liberar conexão da fila durável: {exc_release}"
                    )

    try:
        # O lock advisory é transacional para continuar correto atrás de
        # poolers em modo transaction. A transação externa permanece aberta
        # durante parse + persistência; db_ops usa apenas um savepoint interno.
        transacao_importacao = conn_bg.transaction()
        await transacao_importacao.start()
        lock_acquired = await conn_bg.fetchval(
            "SELECT pg_try_advisory_xact_lock($1)", lock_key
        )
        if not lock_acquired:
            _registrar(
                conciliacao_id,
                status="erro",
                etapa="erro",
                erro_fatal="Já existe uma importação em andamento para este projeto.",
            )
            return

        projeto_ja_populado = await conn_bg.fetchval(
            '''
            select exists (select 1 from transacoes where projeto_id = $1)
                or exists (
                    select 1 from extrato_movimentos m
                    join contas_captadoras c on c.id = m.conta_id
                    where c.projeto_id = $1
                )
            ''',
            projeto_id,
        )
        if projeto_ja_populado:
            raise RuntimeError(
                "O projeto já possui transações ou extratos importados. "
                "A substituição automática foi bloqueada até existir importação versionada."
            )

        base = Path(tempfile.mkdtemp(prefix="importacao_pasta_"))
        _registrar(conciliacao_id, base=str(base))
        _registrar(conciliacao_id, status="em_progresso", etapa="preparando arquivos", progresso=10)
        
        pag_dir = base / "pagamentos"
        pag_dir.mkdir(parents=True, exist_ok=True)
        ext_dir = base / "extratos"
        ext_dir.mkdir(parents=True, exist_ok=True)
        
        if extrato_bytes and extrato_nome:
            extrato_file = ext_dir / Path(extrato_nome).name
            extrato_file.write_bytes(extrato_bytes)
        
        hashes_salvos = set()
        planilha_base_bytes = None
        
        for nome_arq, c_bytes in comprovantes_dados:
            if nome_arq.lower().endswith(".zip"):
                import io
                import zipfile
                with zipfile.ZipFile(io.BytesIO(c_bytes)) as z:
                    for member in z.namelist():
                        if member.startswith("__MACOSX") or member.endswith("/") or Path(member).name.startswith("."):
                            continue
                        member_bytes = z.read(member)
                        
                        m_hash = hashlib.sha256(member_bytes).hexdigest()
                        if m_hash in hashes_salvos:
                            continue
                        hashes_salvos.add(m_hash)
                        
                        caminho_dest = pag_dir / Path(member)
                        caminho_dest.parent.mkdir(parents=True, exist_ok=True)
                        caminho_dest.write_bytes(member_bytes)
            else:
                f_hash = hashlib.sha256(c_bytes).hexdigest()
                if f_hash in hashes_salvos:
                    continue
                hashes_salvos.add(f_hash)
                
                caminho_dest = pag_dir / Path(nome_arq)
                caminho_dest.parent.mkdir(parents=True, exist_ok=True)
                caminho_dest.write_bytes(c_bytes)

        from backend.services.sincronizacao_documentos_service import tipo_planilha
        for p in pag_dir.glob("**/*"):
            if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"}:
                candidato = p.read_bytes()
                if tipo_planilha(candidato) == "planilha_base":
                    planilha_base_bytes = candidato
                    break
        
        # Move arquivos identificados como extrato para a pasta de extratos
        for p in list(pag_dir.glob("**/*")):
            if p.is_file():
                p_name_lower = p.name.lower()
                p_path_lower = str(p).lower()
                if "extrato" in p_name_lower or "extract" in p_name_lower or "conta_corrente" in p_name_lower or "extratos" in p_path_lower or "extrato" in p_path_lower:
                    dest = ext_dir / p.name
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(p), str(dest))
        
        raiz_pag = pag_dir
        
        _registrar(conciliacao_id, etapa="lendo comprovantes", progresso=30)
        comprovantes, sem_valor = await asyncio.to_thread(_parse_comprovantes, raiz_pag)
        
        _registrar(conciliacao_id, etapa="lendo extratos", progresso=50)
        movimentos = await asyncio.to_thread(_parse_extratos, ext_dir)
        
        _registrar(conciliacao_id, etapa="conciliando e cruzando dados", progresso=70)
        resultado = await asyncio.to_thread(conciliar, comprovantes, movimentos)

        def data_iso_ou_none(valor):
            if isinstance(valor, date):
                return valor
            if isinstance(valor, str):
                try:
                    return date.fromisoformat(valor)
                except ValueError:
                    return None
            return None

        for movimento in movimentos:
            if data_iso_ou_none(movimento.get("data")) is None:
                raise RuntimeError(
                    f"Movimento de extrato com data inválida: {movimento.get('data')!r}."
                )
        
        _registrar(conciliacao_id, etapa="gravando dados no banco de dados", progresso=85)
        
        async def db_ops():
            try:
                async with conn_bg.transaction():
                    conta = await conn_bg.fetchrow("select id from contas_captadoras where projeto_id = $1", projeto_id)
                    if not conta:
                        conta = await conn_bg.fetchrow("insert into contas_captadoras (projeto_id) values ($1) returning id", projeto_id)
                    conta_id = conta["id"]
                    
                    transacoes_inseridas = {}
                    
                    for comp in comprovantes:
                        nome_doc = Path(comp["fonte"]).name
                        prestador_extraido = None
                        item_extraido = None
                        
                        match_dashes = re.match(r"^\d+\s*-\s*\d{2}-\d{2}-\d{4}\s*-\s*([^-(\n]+)\s*-\s*([^-.\n]+)", nome_doc)
                        if match_dashes:
                            prestador_extraido = match_dashes.group(1).strip()
                            item_extraido = match_dashes.group(2).replace(".pdf", "").replace(".png", "").replace(".jpg", "").strip()
                        else:
                            match_dot = re.match(r"^\d+\.\s+([^-(\n]+)\s*-\s*([^-.\n]+)", nome_doc)
                            if match_dot:
                                prestador_extraido = match_dot.group(1).strip()
                                item_extraido = match_dot.group(2).replace(".pdf", "").replace(".png", "").replace(".jpg", "").strip()

                        cnpj_cpf = comp.get("cnpj")
                        is_cnpj = False
                        if cnpj_cpf:
                            is_cnpj = len(re.sub(r"\D", "", cnpj_cpf)) == 14

                        prestador = prestador_extraido or comp.get("favorecido") or "Prestador de Serviço"
                        favorecido_banco = comp.get("favorecido") or prestador_extraido
                        razao_social = f"Favorecido no extrato: {favorecido_banco}." if favorecido_banco else "-"

                        data_pag = data_iso_ou_none(comp.get("data"))
                            
                        valor = Decimal(str(comp.get("valor") or 0))
                        tipo_doc = parse_tipo_doc(comp["fonte"]) or "OUTRO"
                        tem_nf = tipo_doc == "NFE"
                        tem_comprovante = bool(comp.get("tem_sisbb"))
                        
                        row_t = await conn_bg.fetchrow(
                            '''
                            insert into transacoes (
                                projeto_id, fornecedor, razao_social, documento, data_pagamento,
                                valor_bruto, valor_liquido, tem_nf, tem_comprovante, status
                            ) values ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                            returning id
                            ''',
                            projeto_id, prestador, razao_social, cnpj_cpf, data_pag,
                            valor, valor, tem_nf, tem_comprovante, "PENDENTE"
                        )
                        t_id = row_t["id"]
                        
                        origem_doc = Path(comp.get("caminho") or comp["fonte"])
                        if not origem_doc.is_file():
                            origem_doc = raiz_pag / Path(comp["fonte"]).name
                        conteudo_doc = origem_doc.read_bytes()
                        hash_doc = hashlib.sha256(conteudo_doc).hexdigest()
                        extensao_doc = origem_doc.suffix.lower() or ".bin"
                        arquivo_ref, criado_nesta_execucao = await asyncio.to_thread(
                            criar_arquivo_se_ausente,
                            f"{projeto_id}/comprovantes/{hash_doc}{extensao_doc}",
                            conteudo_doc,
                        )
                        if criado_nesta_execucao:
                            objetos_criados.append((arquivo_ref, hash_doc))
                        await conn_bg.execute(
                            '''
                            insert into documentos_transacao (transacao_id, tipo, arquivo_ref, confianca_ocr)
                            values ($1, $2, $3, $4)
                            ''',
                            t_id, tipo_doc, arquivo_ref, comp.get("confianca_ocr")
                        )
                        # A mesma transação que persiste a referência invalida
                        # qualquer coleta pendente desta chave.
                        await cancelar_pendencia(conn_bg, arquivo_ref)
                        
                        descricao_despesa = item_extraido or "Serviço Prestado"
                        await conn_bg.execute(
                            '''
                            insert into despesas (transacao_id, projeto_id, descricao, valor)
                            values ($1, $2, $3, $4)
                            ''',
                            t_id, projeto_id, descricao_despesa, valor
                        )
                        
                        transacoes_inseridas[str(comp["fonte"]).lower()] = t_id
                        transacoes_inseridas[Path(comp["fonte"]).name.lower()] = t_id
                    
                    mapa_status_cruzado = {}
                    mapa_transacao_vincular = {}
                    
                    for l in resultado["linhas"]:
                        d_str = l.get("data")
                        d_obj = None
                        if isinstance(d_str, str):
                            try:
                                d_obj = date.fromisoformat(d_str)
                            except ValueError:
                                pass
                        elif isinstance(d_str, date):
                            d_obj = d_str
                        
                        val_dec = None
                        if l.get("valor") is not None:
                            val_dec = abs(Decimal(str(l.get("valor"))))
                            
                        doc_str = str(l.get("doc_extrato") or "")
                        chave_mov = (d_obj, doc_str, val_dec)
                        
                        if l.get("status") == "conferido":
                            mapa_status_cruzado[chave_mov] = "CONCILIADO"
                            fonte = str(l.get("fonte_comprovante") or "").lower()
                            fonte_nome = Path(fonte).name if fonte else ""
                            v_id = transacoes_inseridas.get(fonte) or transacoes_inseridas.get(fonte_nome)
                            if not v_id and d_obj and val_dec:
                                row_fb = await conn_bg.fetchrow(
                                    "select id from transacoes where projeto_id = $1 and data_pagamento = $2 and abs(valor_bruto - $3) < 0.01 limit 1",
                                    projeto_id, d_obj, val_dec
                                )
                                if row_fb:
                                    v_id = row_fb["id"]
                            if v_id:
                                mapa_transacao_vincular[chave_mov] = v_id
                        else:
                            mapa_status_cruzado[chave_mov] = "PENDENTE"
                    
                    for m in movimentos:
                        valor_m = Decimal(str(m["valor"]))
                        if m["sinal"] == "D":
                            valor_m = -valor_m
                            
                        data_mov = data_iso_ou_none(m["data"])
                        
                        doc_mov_str = str(m.get("doc") or "")
                        chave_mov = (data_mov, doc_mov_str, abs(valor_m))
                        status_mov = mapa_status_cruzado.get(chave_mov, "PENDENTE")
                        t_id_vinculo = mapa_transacao_vincular.get(chave_mov, None)
                        
                        row_m = await conn_bg.fetchrow(
                            '''
                            insert into extrato_movimentos (
                                conta_id, data, historico, documento, tipo, valor, status_conciliacao
                            ) values ($1, $2, $3, $4, $5, $6, $7)
                            returning id
                            ''',
                            conta_id, data_mov, m.get("historico") or m.get("favorecido"),
                            m["doc"], tipo_por_sinal(m["sinal"]), valor_m, status_mov
                        )
                        mov_id = row_m["id"]
                        
                        if status_mov == "CONCILIADO" and t_id_vinculo:
                            await conn_bg.execute(
                                "update transacoes set status = 'CONCILIADO_OK' where id = $1",
                                t_id_vinculo
                            )
                            await conn_bg.execute(
                                "insert into conciliacao_extrato (transacao_id, movimento_id) values ($1, $2) on conflict do nothing",
                                t_id_vinculo, mov_id
                            )
                            
                    await conn_bg.execute(
                        '''
                        update transacoes set status = 'CONCILIADO_OK'
                        where id in (select transacao_id from conciliacao_extrato)
                          and projeto_id = $1
                        ''',
                        projeto_id
                    )
                    if planilha_base_bytes:
                        from backend.services.sincronizacao_documentos_service import (
                            extrair_rubricas_planilha_base,
                            extrair_vinculos_rubrica_planilha_base,
                            vincular_despesas_planilha_base,
                        )
                        for codigo, descricao, valor_orcado in extrair_rubricas_planilha_base(planilha_base_bytes):
                            await conn_bg.execute(
                                """
                                insert into rubricas (projeto_id, codigo, descricao, valor_orcado)
                                values ($1, $2, $3, $4)
                                on conflict (projeto_id, codigo) do update
                                  set descricao = excluded.descricao,
                                      valor_orcado = excluded.valor_orcado
                                """,
                                projeto_id, codigo, descricao, valor_orcado,
                            )
                        await vincular_despesas_planilha_base(
                            conn_bg,
                            projeto_id,
                            extrair_vinculos_rubrica_planilha_base(planilha_base_bytes),
                        )
            finally:
                pass

        await db_ops()
        commit_iniciado = True
        await transacao_importacao.commit()
        transacao_importacao = None
        # As referências já estão persistidas: a partir daqui estes objetos
        # não são órfãos e jamais podem entrar na compensação, mesmo se o
        # bookkeeping em memória falhar.
        objetos_criados.clear()
        _registrar(
            conciliacao_id,
            status="sucesso",
            progresso=100,
            etapa="concluído",
            resumo=resultado["resumo"]
        )
    except Exception as e:
        logger.exception("Falha na importação de pasta bg %s", conciliacao_id)
        falhas_remocao: list[tuple[str, str, str]] = []

        if not commit_iniciado:
            # O lock ainda pertence à transação original. Compensar antes do
            # rollback impede que um retry persista/reutilize a chave enquanto
            # esta execução a remove.
            for ref, hash_doc in objetos_criados:
                try:
                    removido = await asyncio.to_thread(remover_arquivo, ref)
                    if not removido:
                        falhas_remocao.append(
                            (
                                ref,
                                hash_doc,
                                "storage retornou remoção não confirmada",
                            )
                        )
                except Exception as exc_rem:
                    falhas_remocao.append((ref, hash_doc, str(exc_rem)))
                    logger.error("Falha ao remover órfão %s: %s", ref, exc_rem)

        # Commit iniciado e exceção é resultado ambíguo: nunca autoriza delete.
        # Primeiro encerra/libera a sessão original; depois uma conexão íntegra
        # readquire o lock e consulta as referências persistidas.
        if transacao_importacao is not None:
            try:
                await transacao_importacao.rollback()
            except Exception as exc_rollback:
                erros_operacionais.append(
                    f"falha no rollback da importação: {exc_rollback}"
                )
            finally:
                transacao_importacao = None

        await liberar_conexao_original()

        if commit_iniciado:
            await reconciliar_objetos_em_conexao_nova(
                [
                    (ref, hash_doc, "resultado de commit ambíguo")
                    for ref, hash_doc in objetos_criados
                ],
                tentar_remocao=True,
            )
        elif falhas_remocao:
            # A remoção já falhou enquanto o lock original estava ativo. Após
            # readquirir o lock, revalida referências antes de enfileirar.
            await reconciliar_objetos_em_conexao_nova(
                falhas_remocao,
                tentar_remocao=False,
            )

        erro_fatal = str(e)
        if erros_operacionais:
            erro_fatal += " | Erro operacional: " + " | ".join(erros_operacionais)
        _registrar(
            conciliacao_id,
            status="erro",
            etapa="erro",
            erro_fatal=erro_fatal,
        )
    finally:
        if base is not None:
            shutil.rmtree(base, ignore_errors=True)
        if transacao_importacao is not None:
            try:
                await transacao_importacao.rollback()
            except Exception as exc_rollback:
                logger.error("Falha no rollback final da importação: %s", exc_rollback)
            finally:
                transacao_importacao = None
        await liberar_conexao_original()
